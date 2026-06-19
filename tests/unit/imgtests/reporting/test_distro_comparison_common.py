"""Unit tests for distro_comparison_common module."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook

from imgtests.reporting.distro_comparison_common import (
    CONFIGURATION_SHEET,
    EXPERIMENT_SHEET,
    ComparisonGroup,
    ExperimentInfo,
    MetricBucket,
    MetricExtractionOptions,
    MetricWorksheetContext,
    SheetRows,
    build_comparison_groups,
    build_configuration_distro_map,
    build_experiment_info_map,
    build_group_by_experiment,
    cell_value,
    collect_metric_buckets_from_row,
    column_index,
    find_metric_columns,
    is_comparison_sheet,
    is_invalid_metric_value,
    is_nonzero_number,
    logical_test_name,
    make_group,
    metric_worksheet_context,
    normalize_distro,
    normalize_experiment_type,
    normalize_header,
    normalize_metric_text,
    normalize_selected_ids,
    resolve_output_path,
    row_has_data,
    sanitize_sheet_name,
    sheet_has_data_rows,
    strip_excel_suffix,
    to_finite_float,
    unique_sheet_name,
)


@pytest.mark.parametrize(
    ("sheet_name", "expected"),
    [
        ("comparison_selection", True),
        ("comparison_index", True),
        ("comparison_data", True),
        ("comparison_charts_1", True),
        ("comparison_charts_2", True),
        ("charts_1", True),
        ("charts_metrics", True),
        ("charts_summary", True),
        ("configuration", False),
        ("experiment", False),
        ("test_data", False),
        ("test_data_1", False),
    ],
)
def test_is_comparison_sheet(sheet_name: str, expected: bool) -> None:
    assert is_comparison_sheet(sheet_name) == expected


class TestBuildConfigurationDistroMap:
    def test_empty_workbook(self) -> None:
        workbook = Workbook()
        result = build_configuration_distro_map(workbook)
        assert result == {}

    def test_missing_required_columns(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        ws.title = CONFIGURATION_SHEET
        ws.append(["other_column", "another_column"])

        assert build_configuration_distro_map(workbook) == {}

    def test_configuration_with_none_values(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        ws.title = CONFIGURATION_SHEET
        ws.append(["configuration_id", "distribution_description"])
        ws.append([None, None])

        assert build_configuration_distro_map(workbook) == {}

    def test_valid_configuration_distro_map(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        ws.title = CONFIGURATION_SHEET
        ws.append(["configuration_id", "distribution_description"])
        ws.append([1, "Poky"])
        ws.append([2, "openSuse"])
        ws.append([3, "yocto"])

        assert build_configuration_distro_map(workbook) == {
            "1": "poky",
            "2": "suse",
            "3": "poky",
        }


class TestBuildExperimentInfoMap:
    def test_empty_workbook(self) -> None:
        workbook = Workbook()
        config_distro: dict[str, str] = {}
        assert build_experiment_info_map(workbook, config_distro) == {}

    def test_missing_required_columns(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        ws.title = EXPERIMENT_SHEET
        ws.append(["other_column", "another_column"])
        config_distro: dict[str, str] = {}

        assert build_experiment_info_map(workbook, config_distro) == {}

    def test_workbook_without_none_values(self) -> None:
        workbook = Workbook()
        config_distro: dict[str, str] = {}
        ws = workbook.active
        ws.title = EXPERIMENT_SHEET
        ws.append(
            [
                "experiment_id",
                "configuration_id",
                "description",
                "type",
                "started_at",
            ],
        )
        ws.append([None, None, "test experiment", "performance", "2024-01-01 10:00:00"])

        assert build_experiment_info_map(workbook, config_distro) == {}

    def test_valid_experiment_info_map(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        ws.title = EXPERIMENT_SHEET
        ws.append(
            [
                "experiment_id",
                "configuration_id",
                "description",
                "type",
                "started_at",
            ],
        )
        ws.append([1, 1, "test experiment", "performance", "2024-01-01 10:00:00"])
        ws.append([2, 2, "another test", "endurance", "2024-01-02 11:00:00"])

        config_distro = {"1": "poky", "2": "suse"}

        assert build_experiment_info_map(workbook, config_distro) == {
            1: ExperimentInfo(
                1,
                "poky",
                1,
                "test experiment",
                "performance",
                "2024-01-01 10:00:00",
            ),
            2: ExperimentInfo(
                2,
                "suse",
                2,
                "another test",
                "endurance",
                "2024-01-02 11:00:00",
            ),
        }


class TestBuildComparisonGroups:
    def test_empty_experiment_info(self) -> None:
        experiment_info: dict[int, ExperimentInfo] = {}
        assert build_comparison_groups(experiment_info, experiment_ids=None) == []

    def test_single_distro_only(self) -> None:
        experiment_info = {
            1: ExperimentInfo(
                experiment_id=1,
                distro="poky",
                configuration_id=1,
                description="test",
                experiment_type="performance",
                started_at="2024-01-01",
            ),
        }
        assert build_comparison_groups(experiment_info, experiment_ids=None) == []

    def test_explicit_pair_experiment_ids(self) -> None:
        experiment_info = {
            10: ExperimentInfo(
                experiment_id=10,
                distro="poky",
                configuration_id=10,
                description="explicit test",
                experiment_type="endurance",
                started_at="2024-02-01",
            ),
            20: ExperimentInfo(
                experiment_id=20,
                distro="suse",
                configuration_id=20,
                description="explicit test",
                experiment_type="endurance",
                started_at="2024-02-01",
            ),
        }
        result = build_comparison_groups(
            experiment_info,
            experiment_ids=["10", "20"],
        )

        assert result == [
            ComparisonGroup(
                order=1,
                group_id="exp_10_20",
                label="explicit test | endurance | exp 10 / poky vs exp 20 / suse",
                short_label="1: p10/s20",
                experiments={"poky": experiment_info[10], "suse": experiment_info[20]},
            ),
        ]

    def test_invalid_experiment_ids(self) -> None:
        experiment_info = {
            1: ExperimentInfo(
                experiment_id=1,
                distro="poky",
                configuration_id=1,
                description="test",
                experiment_type="performance",
                started_at="2024-01-01",
            ),
            2: ExperimentInfo(
                experiment_id=2,
                distro="suse",
                configuration_id=2,
                description="test",
                experiment_type="performance",
                started_at="2024-01-01",
            ),
        }

        assert build_comparison_groups(experiment_info, experiment_ids=["abc", "def"]) == []

    def test_multiple_groups(self) -> None:
        experiment_info = {
            1: ExperimentInfo(
                experiment_id=1,
                distro="poky",
                configuration_id=1,
                description="group1",
                experiment_type="performance",
                started_at="2024-01-01",
            ),
            2: ExperimentInfo(
                experiment_id=2,
                distro="suse",
                configuration_id=2,
                description="group1",
                experiment_type="performance",
                started_at="2024-01-01",
            ),
            3: ExperimentInfo(
                experiment_id=3,
                distro="poky",
                configuration_id=3,
                description="group2",
                experiment_type="endurance",
                started_at="2024-02-01",
            ),
            4: ExperimentInfo(
                experiment_id=4,
                distro="suse",
                configuration_id=4,
                description="group2",
                experiment_type="endurance",
                started_at="2024-02-01",
            ),
        }

        assert build_comparison_groups(experiment_info, experiment_ids=None) == [
            ComparisonGroup(
                order=1,
                group_id="exp_1_2",
                label="group1 | performance | exp 1 / poky vs exp 2 / suse",
                short_label="1: p1/s2",
                experiments={
                    "poky": ExperimentInfo(
                        experiment_id=1,
                        distro="poky",
                        configuration_id=1,
                        description="group1",
                        experiment_type="performance",
                        started_at="2024-01-01",
                    ),
                    "suse": ExperimentInfo(
                        experiment_id=2,
                        distro="suse",
                        configuration_id=2,
                        description="group1",
                        experiment_type="performance",
                        started_at="2024-01-01",
                    ),
                },
            ),
            ComparisonGroup(
                order=2,
                group_id="exp_3_4",
                label="group2 | endurance | exp 3 / poky vs exp 4 / suse",
                short_label="2: p3/s4",
                experiments={
                    "poky": ExperimentInfo(
                        experiment_id=3,
                        distro="poky",
                        configuration_id=3,
                        description="group2",
                        experiment_type="endurance",
                        started_at="2024-02-01",
                    ),
                    "suse": ExperimentInfo(
                        experiment_id=4,
                        distro="suse",
                        configuration_id=4,
                        description="group2",
                        experiment_type="endurance",
                        started_at="2024-02-01",
                    ),
                },
            ),
        ]


def test_make_group() -> None:
    poky_info = ExperimentInfo(
        experiment_id=100,
        distro="poky",
        configuration_id=100,
        description="test description",
        experiment_type="performance",
        started_at="2024-01-01",
    )
    suse_info = ExperimentInfo(
        experiment_id=200,
        distro="suse",
        configuration_id=200,
        description="test description",
        experiment_type="performance",
        started_at="2024-01-01",
    )

    assert make_group(poky_info, suse_info, order=5) == ComparisonGroup(
        5,
        "exp_100_200",
        "test description | performance | exp 100 / poky vs exp 200 / suse",
        "5: p100/s200",
        {"poky": poky_info, "suse": suse_info},
    )


@pytest.mark.parametrize(
    ("input_ids", "expected"),
    [
        (None, set()),
        ([], set()),
        (["1", "2", "3"], {1, 2, 3}),
        (["1", "abc", "3", "xyz"], {1, 3}),
        (["100", "200", "300"], {100, 200, 300}),
    ],
)
def test_normalize_selected_ids(input_ids: list[str] | None, expected: set[int]) -> None:
    assert normalize_selected_ids(input_ids) == expected


def test_build_group_by_experiment() -> None:
    poky_info = ExperimentInfo(
        experiment_id=1,
        distro="poky",
        configuration_id=1,
        description="test",
        experiment_type="performance",
        started_at="2024-01-01",
    )
    suse_info = ExperimentInfo(
        experiment_id=2,
        distro="suse",
        configuration_id=2,
        description="test",
        experiment_type="performance",
        started_at="2024-01-01",
    )

    group = make_group(poky_info, suse_info, order=1)
    groups = [group]

    assert build_group_by_experiment(groups) == {
        "1": group,
        "2": group,
    }


class TestMetricWorksheetContext:
    def test_comparison_sheet_returns_none(self) -> None:
        workbook = Workbook()
        ws = workbook.create_sheet("comparison_data")

        assert metric_worksheet_context(ws, lambda _: [1]) is None

    def test_missing_required_columns(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "test_data"
        ws.append(["other_column"])

        assert metric_worksheet_context(ws, lambda _: [1]) is None

    def test_no_metric_columns(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "test_data"
        ws.append(["configuration_id", "experiment_id"])

        assert metric_worksheet_context(ws, lambda _: []) is None

    def test_valid_worksheet_context(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "test_data"
        headers = [
            "configuration_id",
            "experiment_id",
            "test_name",
            "metric1",
            "metric2",
        ]
        ws.append(headers)
        ws.append([1, 100, "test", 1.5, 2.5])

        def column_selector(_: list[str]) -> list[int]:
            return [3, 4]  # metric1, metric2

        result = metric_worksheet_context(ws, column_selector)

        assert result is not None
        assert result.sheet_name == "test_data"
        assert result.source_sheet == "test_data"
        assert result.headers == headers
        assert result.config_index == 0
        assert result.experiment_index == 1
        assert result.metric_indexes == [3, 4]
        assert result.test_name_index == 2  # noqa: PLR2004
        assert result.tool_name_index is None
        assert result.distribution_index is None


class TestCollectMetricBucketsFromRow:
    def test_row_outside_group(self) -> None:
        row = (1, 999, "test", 1.5)  # experiment_id=999 not in group
        context = MetricWorksheetContext(
            sheet_name="test",
            source_sheet="test",
            rows=iter([]),
            headers=["configuration_id", "experiment_id", "test_name", "metric"],
            config_index=0,
            experiment_index=1,
            metric_indexes=[3],
            test_name_index=2,
            tool_name_index=None,
            util_type_index=None,
            distribution_index=None,
        )
        group_by_experiment: dict[str, ComparisonGroup] = {}
        config_distro: dict[str, str] = {}
        buckets: dict[tuple[str, str, str], MetricBucket] = {}
        options = MetricExtractionOptions(column_selector=lambda _: [3])

        collect_metric_buckets_from_row(
            row,
            context,
            group_by_experiment,
            config_distro,
            buckets,
            options,
        )

        assert group_by_experiment == {}

    def test_row_with_invalid_distro(self) -> None:
        row = (1, 1, "test", 1.5)  # experiment_id=1 in group, but distro is invalid
        context = MetricWorksheetContext(
            sheet_name="test",
            source_sheet="test",
            rows=iter([]),
            headers=["configuration_id", "experiment_id", "test_name", "metric"],
            config_index=0,
            experiment_index=1,
            metric_indexes=[3],
            test_name_index=2,
            tool_name_index=None,
            util_type_index=None,
            distribution_index=2,
        )
        poky_info = ExperimentInfo(
            experiment_id=1,
            distro="poky",
            configuration_id=1,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        suse_info = ExperimentInfo(
            experiment_id=2,
            distro="suse",
            configuration_id=2,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        group = make_group(poky_info, suse_info, order=1)
        group_by_experiment = {"1": group, "2": group}
        config_distro = {"1": "poky", "2": "suse"}
        buckets: dict[tuple[str, str, str], MetricBucket] = {}
        options = MetricExtractionOptions(column_selector=lambda _: [3])

        collect_metric_buckets_from_row(
            row,
            context,
            group_by_experiment,
            config_distro,
            buckets,
            options,
        )

        assert buckets == {}

    def test_row_with_invalid_metric_value(self) -> None:
        row = (1, 1, "test", -1.0)  # -1.0 is invalid metric value
        context = MetricWorksheetContext(
            sheet_name="test",
            source_sheet="test",
            rows=iter([]),
            headers=["configuration_id", "experiment_id", "test_name", "metric"],
            config_index=0,
            experiment_index=1,
            metric_indexes=[3],
            test_name_index=2,
            tool_name_index=None,
            util_type_index=None,
            distribution_index=None,
        )
        poky_info = ExperimentInfo(
            experiment_id=1,
            distro="poky",
            configuration_id=1,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        suse_info = ExperimentInfo(
            experiment_id=2,
            distro="suse",
            configuration_id=2,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        group = make_group(poky_info, suse_info, order=1)
        group_by_experiment = {"1": group, "2": group}
        config_distro = {"1": "poky", "2": "suse"}
        buckets: dict[tuple[str, str, str], MetricBucket] = {}
        options = MetricExtractionOptions(column_selector=lambda _: [3])

        collect_metric_buckets_from_row(
            row,
            context,
            group_by_experiment,
            config_distro,
            buckets,
            options,
        )

        assert buckets == {}

    def test_row_with_zero_metric_value_when_require_nonzero(self) -> None:
        row = (1, 1, "test", 0.0)
        context = MetricWorksheetContext(
            sheet_name="test",
            source_sheet="test",
            rows=iter([]),
            headers=["configuration_id", "experiment_id", "test_name", "metric"],
            config_index=0,
            experiment_index=1,
            metric_indexes=[3],
            test_name_index=2,
            tool_name_index=None,
            util_type_index=None,
            distribution_index=None,
        )
        poky_info = ExperimentInfo(
            experiment_id=1,
            distro="poky",
            configuration_id=1,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        suse_info = ExperimentInfo(
            experiment_id=2,
            distro="suse",
            configuration_id=2,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        group = make_group(poky_info, suse_info, order=1)
        group_by_experiment = {"1": group, "2": group}
        config_distro = {"1": "poky", "2": "suse"}
        buckets: dict[tuple[str, str, str], MetricBucket] = {}
        options = MetricExtractionOptions(column_selector=lambda _: [3], require_nonzero=True)

        collect_metric_buckets_from_row(
            row,
            context,
            group_by_experiment,
            config_distro,
            buckets,
            options,
        )

        assert buckets == {}

    def test_valid_row_single_metric(self) -> None:
        row = (1, 1, "test_name", 1.5)
        context = MetricWorksheetContext(
            sheet_name="test",
            source_sheet="test",
            rows=iter([]),
            headers=["configuration_id", "experiment_id", "test_name", "metric"],
            config_index=0,
            experiment_index=1,
            metric_indexes=[3],
            test_name_index=2,
            tool_name_index=None,
            util_type_index=None,
            distribution_index=None,
        )
        poky_info = ExperimentInfo(
            experiment_id=1,
            distro="poky",
            configuration_id=1,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        suse_info = ExperimentInfo(
            experiment_id=2,
            distro="suse",
            configuration_id=2,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        group = make_group(poky_info, suse_info, order=1)
        group_by_experiment = {"1": group, "2": group}
        config_distro = {"1": "poky", "2": "suse"}
        buckets: dict[tuple[str, str, str], MetricBucket] = {}
        options = MetricExtractionOptions(column_selector=lambda _: [3])

        collect_metric_buckets_from_row(
            row,
            context,
            group_by_experiment,
            config_distro,
            buckets,
            options,
        )

        assert len(buckets) == 1
        bucket_key = ("test_name", "test", "metric")
        assert bucket_key in buckets
        bucket = buckets[bucket_key]
        assert bucket.logical_test == "test_name"
        assert bucket.source_sheet == "test"
        assert bucket.metric_name == "metric"
        assert bucket.values[group.group_id]["poky"] == [1.5]
        assert bucket.values[group.group_id]["suse"] == []

    def test_valid_row_multiple_metrics(self) -> None:
        row = (1, 1, "test_name", 1.5, 2.5, 3.5)
        context = MetricWorksheetContext(
            sheet_name="test",
            source_sheet="test",
            rows=iter([]),
            headers=[
                "configuration_id",
                "experiment_id",
                "test_name",
                "metric1",
                "metric2",
                "metric3",
            ],
            config_index=0,
            experiment_index=1,
            metric_indexes=[3, 4, 5],
            test_name_index=2,
            tool_name_index=None,
            util_type_index=None,
            distribution_index=None,
        )
        poky_info = ExperimentInfo(
            experiment_id=1,
            distro="poky",
            configuration_id=1,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        suse_info = ExperimentInfo(
            experiment_id=2,
            distro="suse",
            configuration_id=2,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        group = make_group(poky_info, suse_info, order=1)
        group_by_experiment = {"1": group, "2": group}
        config_distro = {"1": "poky", "2": "suse"}
        buckets: dict[tuple[str, str, str], MetricBucket] = {}
        options = MetricExtractionOptions(column_selector=lambda _: [3, 4, 5])

        collect_metric_buckets_from_row(
            row,
            context,
            group_by_experiment,
            config_distro,
            buckets,
            options,
        )

        assert len(buckets) == 3  # noqa: PLR2004
        assert ("test_name", "test", "metric1") in buckets
        assert ("test_name", "test", "metric2") in buckets
        assert ("test_name", "test", "metric3") in buckets

    def test_row_with_name_mapper(self) -> None:
        row = (1, 1, "test_name", 1.5)
        context = MetricWorksheetContext(
            sheet_name="test",
            source_sheet="test",
            rows=iter([]),
            headers=["configuration_id", "experiment_id", "test_name", "metric"],
            config_index=0,
            experiment_index=1,
            metric_indexes=[3],
            test_name_index=2,
            tool_name_index=None,
            util_type_index=None,
            distribution_index=None,
        )
        poky_info = ExperimentInfo(
            experiment_id=1,
            distro="poky",
            configuration_id=1,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        suse_info = ExperimentInfo(
            experiment_id=2,
            distro="suse",
            configuration_id=2,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        group = make_group(poky_info, suse_info, order=1)
        group_by_experiment = {"1": group, "2": group}
        config_distro = {"1": "poky", "2": "suse"}
        buckets: dict[tuple[str, str, str], MetricBucket] = {}

        def name_mapper(_: str, __: str, metric_name: str) -> str:
            return f"mapped_{metric_name}"

        options = MetricExtractionOptions(
            column_selector=lambda _: [3],
            name_mapper=name_mapper,
        )

        collect_metric_buckets_from_row(
            row,
            context,
            group_by_experiment,
            config_distro,
            buckets,
            options,
        )

        assert len(buckets) == 1
        bucket_key = ("test_name", "test", "mapped_metric")
        assert bucket_key in buckets

    def test_row_with_predicate(self) -> None:
        row = (1, 1, "test_name", 1.5)
        context = MetricWorksheetContext(
            sheet_name="test",
            source_sheet="test",
            rows=iter([]),
            headers=["configuration_id", "experiment_id", "test_name", "metric"],
            config_index=0,
            experiment_index=1,
            metric_indexes=[3],
            test_name_index=2,
            tool_name_index=None,
            util_type_index=None,
            distribution_index=None,
        )
        poky_info = ExperimentInfo(
            experiment_id=1,
            distro="poky",
            configuration_id=1,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        suse_info = ExperimentInfo(
            experiment_id=2,
            distro="suse",
            configuration_id=2,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        group = make_group(poky_info, suse_info, order=1)
        group_by_experiment = {"1": group, "2": group}
        config_distro = {"1": "poky", "2": "suse"}
        buckets: dict[tuple[str, str, str], MetricBucket] = {}

        def predicate(_: str, __: str, metric_name: str) -> bool:
            return metric_name == "metric"

        options = MetricExtractionOptions(
            column_selector=lambda _: [3],
            predicate=predicate,
        )

        collect_metric_buckets_from_row(
            row,
            context,
            group_by_experiment,
            config_distro,
            buckets,
            options,
        )

        assert len(buckets) == 1
        bucket_key = ("test_name", "test", "metric")
        assert bucket_key in buckets

    def test_row_with_predicate_filtered_out(self) -> None:
        row = (1, 1, "test_name", 1.5)
        context = MetricWorksheetContext(
            sheet_name="test",
            source_sheet="test",
            rows=iter([]),
            headers=["configuration_id", "experiment_id", "test_name", "metric"],
            config_index=0,
            experiment_index=1,
            metric_indexes=[3],
            test_name_index=2,
            tool_name_index=None,
            util_type_index=None,
            distribution_index=None,
        )
        poky_info = ExperimentInfo(
            experiment_id=1,
            distro="poky",
            configuration_id=1,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        suse_info = ExperimentInfo(
            experiment_id=2,
            distro="suse",
            configuration_id=2,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        group = make_group(poky_info, suse_info, order=1)
        group_by_experiment = {"1": group, "2": group}
        config_distro = {"1": "poky", "2": "suse"}
        buckets: dict[tuple[str, str, str], MetricBucket] = {}

        def predicate(_: str, __: str, metric_name: str) -> bool:
            return metric_name == "other_metric"

        options = MetricExtractionOptions(
            column_selector=lambda _: [3],
            predicate=predicate,
        )

        collect_metric_buckets_from_row(
            row,
            context,
            group_by_experiment,
            config_distro,
            buckets,
            options,
        )

        assert buckets == {}

    def test_row_with_distribution_index(self) -> None:
        row = (1, 1, "test_name", "poky", 1.5)
        context = MetricWorksheetContext(
            sheet_name="test",
            source_sheet="test",
            rows=iter([]),
            headers=[
                "configuration_id",
                "experiment_id",
                "test_name",
                "distribution_description",
                "metric",
            ],
            config_index=0,
            experiment_index=1,
            metric_indexes=[4],
            test_name_index=2,
            tool_name_index=None,
            util_type_index=None,
            distribution_index=3,
        )
        poky_info = ExperimentInfo(
            experiment_id=1,
            distro="poky",
            configuration_id=1,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        suse_info = ExperimentInfo(
            experiment_id=2,
            distro="suse",
            configuration_id=2,
            description="test",
            experiment_type="performance",
            started_at="2024-01-01",
        )
        group = make_group(poky_info, suse_info, order=1)
        group_by_experiment = {"1": group, "2": group}
        config_distro = {"1": "poky", "2": "suse"}
        buckets: dict[tuple[str, str, str], MetricBucket] = {}
        options = MetricExtractionOptions(column_selector=lambda _: [4])

        collect_metric_buckets_from_row(
            row,
            context,
            group_by_experiment,
            config_distro,
            buckets,
            options,
        )

        assert len(buckets) == 1
        bucket_key = ("test_name", "test", "metric")
        assert bucket_key in buckets
        bucket = buckets[bucket_key]
        assert bucket.values[group.group_id]["poky"] == [1.5]


@pytest.mark.parametrize(
    ("headers", "expected"),
    [
        ([], []),
        (
            ["configuration_id", "experiment_id", "test_name", "metric_value"],
            [3],
        ),
        (["configuration_id", "experiment_id", "value"], [2]),
        (["id", "pid", "uid", "value", "___"], [3]),
        (["test_name", "cpu_usage", "memory_mb", "elapsed_time"], [1, 2, 3]),
    ],
)
def test_find_metric_columns(headers: list[str], expected: list[int]) -> None:
    assert find_metric_columns(headers) == expected


@pytest.mark.parametrize(
    (
        "sheet_name",
        "test_name",
        "tool_name",
        "util_type",
        "expected",
    ),
    [
        (
            "test_sheet",
            "test_name",
            "tool_name",
            "util_type",
            "test_name / tool_name / util_type",
        ),
        (
            "test_sheet_1",
            "",
            "tool_name",
            "",
            "test_sheet / tool_name",
        ),
        (
            "test_sheet",
            "my_tool_test",
            "my_tool",
            "",
            "my_tool_test",
        ),
        (
            "test_sheet_122",
            "",
            "",
            "",
            "test_sheet",
        ),
    ],
)
def test_logical_test_name(
    sheet_name: str,
    test_name: str,
    tool_name: str,
    util_type: str,
    expected: str,
) -> None:
    assert logical_test_name(sheet_name, test_name, tool_name, util_type) == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("test_sheet_1", "test_sheet"),
        ("test_sheet_123", "test_sheet"),
        ("test_sheet", "test_sheet"),
        ("", ""),
    ],
)
def test_strip_excel_suffix(value: str, expected: str) -> None:
    assert strip_excel_suffix(value) == expected


class TestNormalizeHeader:
    def test_normal_string(self) -> None:
        assert normalize_header("test_header") == "test_header"

    def test_none_value(self) -> None:
        assert normalize_header(None) == ""

    def test_whitespace(self) -> None:
        assert normalize_header("  test_header  ") == "test_header"


class TestColumnIndex:
    def test_existing_column(self) -> None:
        headers = ["a", "b", "c", "d"]
        assert column_index(headers, "c") == 2  # noqa: PLR2004

    def test_missing_column(self) -> None:
        headers = ["a", "b", "c"]
        assert column_index(headers, "x") is None

    def test_empty_headers(self) -> None:
        assert column_index([], "a") is None


class TestCellValue:
    def test_valid_index(self) -> None:
        row = (1, 2, 3, 4)
        assert cell_value(row, 2) == 3  # noqa: PLR2004

    def test_index_out_of_bounds(self) -> None:
        row = (1, 2, 3)
        assert cell_value(row, 10) is None

    def test_none_index(self) -> None:
        row = (1, 2, 3)
        assert cell_value(row, None) is None


class TestRowHasData:
    def test_row_with_data(self) -> None:
        assert row_has_data((1, 2, 3)) is True

    def test_row_with_none_and_empty(self) -> None:
        assert row_has_data((None, "", None)) is False

    def test_empty_row(self) -> None:
        assert row_has_data(()) is False


class TestSheetHasDataRows:
    def test_sheet_with_data_rows(self) -> None:
        sheet = SheetRows(
            name="test",
            headers=["a", "b"],
            rows=[[1, 2], [None, None]],
        )
        assert sheet_has_data_rows(sheet) is True

    def test_sheet_without_data_rows(self) -> None:
        sheet = SheetRows(
            name="test",
            headers=["a", "b"],
            rows=[[None, None], ["", ""]],
        )
        assert sheet_has_data_rows(sheet) is False


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("poky", "poky"),
        ("POKY", "poky"),
        ("yocto", "poky"),
        ("suse", "suse"),
        ("SUSE", "suse"),
        ("ubuntu", "ubuntu"),
        ("", ""),
    ],
)
def test_normalize_distro(value: str, expected: str) -> None:
    assert normalize_distro(value) == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("cpu_usage", "cpu_usage"),
        ("cpu-usage %", "cpu_usage"),
        ("memory (MB)", "memory_mb"),
        ("cpu___usage", "cpu_usage"),
    ],
)
def test_normalize_metric_text(value: str, expected: str) -> None:
    assert normalize_metric_text(value) == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("performance", "performance"),
        ("endurance", "endurance"),
        ("invalid", ""),
        (None, ""),
    ],
)
def test_normalize_experiment_type(value: str | None, expected: str) -> None:
    assert normalize_experiment_type(value) == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (True, None),
        (False, None),
        (1, 1.0),
        (1.5, 1.5),
        (float("inf"), None),
        (float("-inf"), None),
        (float("nan"), None),
        ("", None),
        ("abc", None),
        ("1.5", 1.5),
        ("1,5", 1.5),
        ("  1.5  ", 1.5),
        ((1.0,), None),
    ],
)
def test_to_finite_float(value: Any, expected: float | None) -> None:
    assert to_finite_float(value) == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (1.0, True),
        (-1.0, True),
        (0.0, False),
        (1e-13, False),
        (-1e-13, False),
        (None, False),
    ],
)
def test_is_nonzero_number(value: float | None, expected: bool) -> None:
    assert is_nonzero_number(value) == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (-1.0, True),
        (-1, True),
        (0.0, False),
        (1.0, False),
        (None, False),
    ],
)
def test_is_invalid_metric_value(value: float | None, expected: bool) -> None:
    assert is_invalid_metric_value(value) == expected


@pytest.mark.parametrize(
    ("input_path", "output_path", "suffix", "expected"),
    [
        (
            Path("/test/input.xlsx"),
            None,
            "output",
            Path("/test/input_output.xlsx"),
        ),
        (
            Path("/test/input.xlsx"),
            Path("/output/output.xlsx"),
            "output",
            Path("/output/output.xlsx"),
        ),
        (
            Path("/test/input.xlsx"),
            Path("/output"),
            "output",
            Path("/output/input_output.xlsx"),
        ),
    ],
)
def test_resolve_output_path(
    input_path: Path,
    output_path: Path | None,
    suffix: str,
    expected: Path,
) -> None:
    assert resolve_output_path(input_path, output_path, suffix) == expected


@pytest.mark.parametrize(
    ("sheet_name", "used_names", "expected"),
    [
        ("new_sheet", set(), "new_sheet"),
        ("test_sheet", {"test_sheet", "test_sheet_1"}, "test_sheet_2"),
        (
            "a" * 40,
            set(),
            "a" * 31,
        ),
    ],
)
def test_unique_sheet_name(
    sheet_name: str,
    used_names: set[str],
    expected: str,
) -> None:
    result = unique_sheet_name(sheet_name, used_names)

    assert result == expected
    assert result in used_names


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("valid_name", "valid_name"),
        ("name:with?special*chars", "name_with_special_chars"),
        ("name[with]brackets", "name_with_brackets"),
        ("", "Sheet"),
        ("'quoted'", "quoted"),
    ],
)
def test_sanitize_sheet_name(value: str, expected: str) -> None:
    assert sanitize_sheet_name(value) == expected
