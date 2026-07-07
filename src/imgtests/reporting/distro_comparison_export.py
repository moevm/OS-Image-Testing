from __future__ import annotations

import logging
import math
import re
import statistics
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any, Final, Literal

from openpyxl import Workbook, load_workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, Reference, StockChart
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.updown_bars import UpDownBars
from openpyxl.styles import Alignment, Font, PatternFill

from imgtests.reporting.distro_comparison_common import (
    CHART_SHEET_PREFIX,
    COMPARISON_DATA_SHEET,
    COMPARISON_INDEX_SHEET,
    COMPARISON_SELECTION_SHEET,
    COUNT_PATTERNS,
    DISTROS,
    IPERF_PATTERNS,
    MAX_SHEET_NAME_LENGTH,
    MEMORY_PATTERNS,
    PERCENT_PATTERNS,
    STRESS_NG_PATTERNS,
    THROUGHPUT_PATTERNS,
    TIME_PATTERNS,
    ComparisonGroup,
    MetricBucket,
    MetricExtractionOptions,
    SheetRows,
    build_comparison_groups,
    build_configuration_distro_map,
    build_experiment_info_map,
    find_comparison_metric_columns,
    is_comparison_sheet,
    is_ignored_metric_name,
    is_nonzero_number,
    metric_has_any_pattern,
    metric_matches,
    normalize_header,
    normalize_metric_text,
    row_has_data,
    sanitize_sheet_name,
    sheet_has_data_rows,
    style_table,
    unique_sheet_name,
    write_matrix,
)
from imgtests.reporting.distro_comparison_common import (
    extract_metric_buckets as extract_common_metric_buckets,
)
from imgtests.reporting.distro_comparison_common import (
    resolve_output_path as resolve_common_output_path,
)

if TYPE_CHECKING:
    from collections.abc import Sequence
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

type MetricDirection = Literal["higher", "lower", "context_dependent"]

COMPARISON_OUTPUT_NAME: Final = "comparison"
COMPARISON_METRICS_SHEET: Final = "comparison_metrics"

DEFAULT_MAX_CHARTS: Final = 0
DEFAULT_CHARTS_PER_SHEET: Final = 30
MIN_TREND_POINTS: Final = 3
FIRST_DATA_ROW: Final = 2
DUPLICATE_METRIC_SUFFIX_RE: Final = re.compile(r"^(?P<base>.+)_(?P<index>[2-9]|[1-9]\d+)$")
UNSPECIFIED_UNIT: Final = "not specified"
VARIABILITY_METRIC_SUFFIXES: Final = ("_dev", "_stddev")

ISSUE_COUNT_PATTERNS: Final = {
    "broken",
    "error",
    "errors",
    "failed",
    "failure",
    "failures",
    "lost",
    "loss",
    "oom",
    "packet_loss",
    "retransmit",
    "retransmits",
    "skipped",
    "untrustworthy",
}

CONTEXT_DEPENDENT_PATTERNS: Final = {
    "cpu",
    "cpu_used",
    "disk_usage",
    "memory",
    "mem",
    "ram",
    "resource",
    "rss",
    "status",
    "usage",
    "util",
}

FIO_PATTERNS: Final = {
    "fio",
}

NETWORK_VOLUME_PATTERNS: Final = {
    "bytes",
    "packets",
}

WORK_COUNT_PATTERNS: Final = {
    "bogo_ops",
    "operations",
    "ops",
    "total_ios",
}

DATA_VOLUME_PATTERNS: Final = {
    "bytes",
    "kbytes",
}

CHART_SCORE_RULES: tuple[tuple[str, int], ...] = (
    ("iperf", 140),
    ("network_loopback", 140),
    ("systemd", 130),
    ("bitrate", 120),
    ("bits_per_second", 120),
    ("bps", 120),
    ("bandwidth", 115),
    ("iops", 110),
    ("ops_per_sec", 105),
    ("ops_s", 105),
    ("bogo_ops", 100),
    ("pps", 95),
    ("latency", 90),
    ("clat", 88),
    ("slat", 86),
    ("rtt", 84),
    ("duration", 80),
    ("elapsed", 78),
    ("runtime", 76),
    ("seconds", 74),
    ("time", 70),
    ("rss", 65),
    ("memory", 65),
    ("retransmit", 60),
)


@dataclass(frozen=True)
class DistributionStats:
    mean: float
    low: float
    q1: float
    q3: float
    high: float
    samples: int


@dataclass(frozen=True)
class ChartPoint:
    category: str
    poky: float | None
    suse: float | None
    poky_stats: DistributionStats | None = None
    suse_stats: DistributionStats | None = None


@dataclass(frozen=True)
class ChartSpec:
    chart_no: int
    logical_test: str
    source_sheet: str
    metric_name: str
    chart_kind: str
    y_axis_title: str
    reason: str
    points: tuple[ChartPoint, ...]


@dataclass(frozen=True)
class MetricDefinition:
    metric_name: str
    base_metric_name: str
    metric_kind: str
    unit: str
    description: str
    preferred_direction: MetricDirection
    higher_is_better: bool | None


@dataclass(frozen=True)
class DistroComparisonExportOptions:
    output_path: Path | None = None
    experiment_ids: Sequence[str] | None = None
    max_charts: int = DEFAULT_MAX_CHARTS
    charts_per_sheet: int = DEFAULT_CHARTS_PER_SHEET
    copy_source_sheets: bool = False
    include_comparison: bool = True


def build_report(
    input_path: Path,
    *,
    options: DistroComparisonExportOptions,
) -> Path:
    source_wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        source_sheets = read_source_sheets(source_wb) if should_copy_source_sheets(options) else []
        config_distro = build_configuration_distro_map(source_wb)
        experiment_info = build_experiment_info_map(source_wb, config_distro)

        groups: list[ComparisonGroup] = []
        metric_definitions: list[MetricDefinition] = []
        chart_specs: list[ChartSpec] = []

        if options.include_comparison:
            groups = build_comparison_groups(
                experiment_info,
                experiment_ids=(
                    list(options.experiment_ids) if options.experiment_ids is not None else None
                ),
            )
            logger.info("Comparable Poky/SUSE groups: %s", len(groups))
            if groups:
                buckets = extract_metric_buckets(source_wb, config_distro, groups)
                logger.info("Metric buckets with data: %s", len(buckets))
                metric_definitions = build_metric_definitions(buckets)
                chart_specs = build_chart_specs(buckets, groups, max_charts=options.max_charts)
                logger.info("Charts built: %s", len(chart_specs))
    finally:
        source_wb.close()

    result_wb = Workbook()
    result_wb.remove(result_wb.active)
    used_sheet_names: set[str] = set()

    if options.include_comparison:
        write_comparison_selection(result_wb, used_sheet_names, groups, len(chart_specs))
        write_comparison_metrics(result_wb, used_sheet_names, metric_definitions)
        write_comparison_index(result_wb, used_sheet_names, chart_specs)
        write_comparison_data(result_wb, used_sheet_names, chart_specs)
        write_comparison_charts(
            result_wb,
            used_sheet_names,
            chart_specs,
            charts_per_sheet=options.charts_per_sheet,
        )

    for sheet in source_sheets:
        if not sheet_has_data_rows(sheet):
            continue
        ws = result_wb.create_sheet(unique_sheet_name(sheet.name, used_sheet_names))
        write_table_sheet(ws, sheet)

    if not result_wb.worksheets:
        ws = result_wb.create_sheet("no_data")
        ws["A1"] = "No data found."

    result_path = resolve_common_output_path(
        input_path,
        options.output_path,
        COMPARISON_OUTPUT_NAME,
    )
    result_path.parent.mkdir(parents=True, exist_ok=True)
    result_wb.save(result_path)
    return result_path


def should_copy_source_sheets(options: DistroComparisonExportOptions) -> bool:
    return options.copy_source_sheets or not options.include_comparison


def export_distro_comparison_to_excel(
    input_path: Path,
    options: DistroComparisonExportOptions | None = None,
) -> Path:
    options = options or DistroComparisonExportOptions()
    return build_report(
        input_path=input_path,
        options=options,
    )


def collect_metric_definitions(
    input_path: Path,
    *,
    experiment_ids: Sequence[str] | None = None,
) -> list[MetricDefinition]:
    source_wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        config_distro = build_configuration_distro_map(source_wb)
        experiment_info = build_experiment_info_map(source_wb, config_distro)
        groups = build_comparison_groups(
            experiment_info,
            experiment_ids=list(experiment_ids) if experiment_ids is not None else None,
        )
        if not groups:
            return []
        buckets = extract_metric_buckets(source_wb, config_distro, groups)
        return build_metric_definitions(buckets)
    finally:
        source_wb.close()


def read_source_sheets(workbook: Workbook) -> list[SheetRows]:
    sheets: list[SheetRows] = []
    for worksheet in workbook.worksheets:
        if is_comparison_sheet(worksheet.title):
            continue
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            headers = [normalize_header(value) for value in next(rows_iter)]
        except StopIteration:
            continue
        rows = [list(row) for row in rows_iter if row_has_data(row)]
        sheets.append(SheetRows(name=worksheet.title, headers=headers, rows=rows))
    return sheets


def extract_metric_buckets(
    workbook: Any,
    config_distro: dict[str, str],
    groups: list[ComparisonGroup],
) -> list[MetricBucket]:
    return extract_common_metric_buckets(
        workbook,
        config_distro,
        groups,
        MetricExtractionOptions(column_selector=find_comparison_metric_columns),
    )


def build_metric_definitions(buckets: list[MetricBucket]) -> list[MetricDefinition]:
    definitions: dict[str, MetricDefinition] = {}

    for bucket in buckets:
        if is_ignored_metric_name(bucket.metric_name):
            continue
        definition = describe_metric_bucket(bucket)
        if not should_report_metric_definition(definition):
            continue
        definition_key = definition.base_metric_name
        existing_definition = definitions.get(definition_key)
        if existing_definition is None or metric_definition_score(
            definition,
        ) > metric_definition_score(existing_definition):
            definitions[definition_key] = definition

    return [definitions[key] for key in sorted(definitions)]


def describe_metric_bucket(bucket: MetricBucket) -> MetricDefinition:
    metric_name = normalize_metric_text(bucket.metric_name)
    base_metric_name = base_metric_name_for(metric_name)
    metric_text = normalize_metric_text(base_metric_name)
    metric_tokens = set(metric_text.split("_"))
    full_text = normalize_metric_text(
        f"{bucket.logical_test} {bucket.source_sheet} {base_metric_name}",
    )
    full_tokens = set(full_text.split("_"))

    metric_kind = metric_kind_for(bucket, metric_text, metric_tokens)

    if metric_kind == "generic" and metric_has_any_pattern(full_text, full_tokens, TIME_PATTERNS):
        metric_kind = "time"

    unit = metric_unit_for(metric_text, metric_tokens, metric_kind)
    direction, higher_is_better = metric_direction_for(metric_kind, metric_text, metric_tokens)

    return MetricDefinition(
        metric_name=metric_name,
        base_metric_name=base_metric_name,
        metric_kind=metric_kind,
        unit=unit,
        description=metric_description(base_metric_name, metric_kind, unit),
        preferred_direction=direction,
        higher_is_better=higher_is_better,
    )


def should_report_metric_definition(definition: MetricDefinition) -> bool:
    return definition.metric_kind != "generic"


def base_metric_name_for(metric_name: str) -> str:
    normalized = normalize_metric_text(metric_name)
    match = DUPLICATE_METRIC_SUFFIX_RE.match(normalized)
    if match is None:
        return normalized

    base_name = match.group("base")
    base_tokens = set(base_name.split("_"))
    if is_known_metric_name(base_name, base_tokens):
        return base_name

    return normalized


def is_known_metric_name(metric_text: str, metric_tokens: set[str]) -> bool:
    return is_variability_metric(metric_text) or metric_has_any_pattern(
        metric_text,
        metric_tokens,
        (
            THROUGHPUT_PATTERNS
            | TIME_PATTERNS
            | ISSUE_COUNT_PATTERNS
            | NETWORK_VOLUME_PATTERNS
            | MEMORY_PATTERNS
            | PERCENT_PATTERNS
            | WORK_COUNT_PATTERNS
            | COUNT_PATTERNS
            | DATA_VOLUME_PATTERNS
        ),
    )


def metric_definition_score(definition: MetricDefinition) -> tuple[int, int, int, int]:
    return (
        int(definition.metric_kind != "generic"),
        int(definition.metric_name == definition.base_metric_name),
        int(definition.higher_is_better is not None),
        int(bool(definition.unit)),
    )


def metric_kind_for(  # noqa: C901, PLR0911, PLR0912
    bucket: MetricBucket,
    metric_text: str,
    metric_tokens: set[str],
) -> str:
    if is_variability_metric(metric_text):
        return "variation"
    if is_transfer_volume_metric(metric_text, metric_tokens):
        return "data_volume"
    if metric_has_any_pattern(metric_text, metric_tokens, THROUGHPUT_PATTERNS):
        return "throughput"
    if is_iperf_interval_time_metric(bucket, metric_text, metric_tokens):
        return "interval_time"
    if metric_has_any_pattern(metric_text, metric_tokens, TIME_PATTERNS):
        return "time"
    if is_reliability_percent_metric(metric_text, metric_tokens):
        return "reliability_percent"
    if metric_has_any_pattern(metric_text, metric_tokens, ISSUE_COUNT_PATTERNS):
        return "reliability_count"
    if is_network_volume_metric(bucket, metric_text, metric_tokens):
        return "network_volume"
    if is_io_volume_metric(bucket, metric_text, metric_tokens):
        return "io_volume"
    if metric_has_any_pattern(metric_text, metric_tokens, MEMORY_PATTERNS):
        return "memory"
    if metric_has_any_pattern(metric_text, metric_tokens, PERCENT_PATTERNS):
        return "utilization"
    if metric_has_any_pattern(metric_text, metric_tokens, WORK_COUNT_PATTERNS):
        return "work_count"
    if metric_has_any_pattern(metric_text, metric_tokens, COUNT_PATTERNS):
        return "count"
    if metric_has_any_pattern(metric_text, metric_tokens, DATA_VOLUME_PATTERNS):
        return "data_volume"
    return "generic"


def is_variability_metric(metric_text: str) -> bool:
    normalized = normalize_metric_text(metric_text)
    return normalized.endswith(VARIABILITY_METRIC_SUFFIXES)


def is_transfer_volume_metric(metric_text: str, metric_tokens: set[str]) -> bool:
    return metric_has_any_pattern(
        metric_text,
        metric_tokens,
        {"transfer", "transferred"},
    ) and metric_has_any_pattern(
        metric_text,
        metric_tokens,
        {"bytes", "kbytes", "kb", "mb", "gb"},
    )


def is_reliability_percent_metric(metric_text: str, metric_tokens: set[str]) -> bool:
    return metric_has_any_pattern(
        metric_text,
        metric_tokens,
        ISSUE_COUNT_PATTERNS,
    ) and metric_has_any_pattern(metric_text, metric_tokens, PERCENT_PATTERNS)


def is_network_volume_metric(
    bucket: MetricBucket,
    metric_text: str,
    metric_tokens: set[str],
) -> bool:
    return is_iperf_metric(bucket, metric_text, metric_tokens) and metric_has_any_pattern(
        metric_text,
        metric_tokens,
        NETWORK_VOLUME_PATTERNS,
    )


def is_io_volume_metric(
    bucket: MetricBucket,
    metric_text: str,
    metric_tokens: set[str],
) -> bool:
    if not metric_has_any_pattern(metric_text, metric_tokens, DATA_VOLUME_PATTERNS):
        return False
    return is_fio_test_bucket(bucket) or metric_has_any_pattern(
        metric_text,
        metric_tokens,
        {
            "io_bytes",
            "io_kbytes",
            "read_bytes",
            "read_io_bytes",
            "read_io_kbytes",
            "write_bytes",
            "write_io_bytes",
            "write_io_kbytes",
        },
    )


def is_iperf_interval_time_metric(
    bucket: MetricBucket,
    metric_text: str,
    metric_tokens: set[str],
) -> bool:
    return is_iperf_metric(bucket, metric_text, metric_tokens) and is_interval_time_text(
        metric_text,
        metric_tokens,
    )


def is_interval_time_text(metric_text: str, metric_tokens: set[str]) -> bool:
    return metric_has_any_pattern(
        metric_text,
        metric_tokens,
        {"sec", "secs", "seconds"},
    ) and not metric_has_any_pattern(metric_text, metric_tokens, THROUGHPUT_PATTERNS)


def metric_unit_for(metric_text: str, metric_tokens: set[str], metric_kind: str) -> str:
    unit = ""
    rules = (
        ({"pps"}, "packets/s"),
        ({"iops"}, "IOPS"),
        ({"microseconds_per_op", "usec_per_op", "usecs_per_op"}, "us/op"),
        (
            {
                "bogo_ops_per_sec",
                "bogo_ops_s",
                "ops_per_sec",
                "ops_s",
                "requests_per_second",
            },
            "ops/s",
        ),
        ({"records_per_second"}, "records/s"),
        ({"gb_per_sec"}, "GB/s"),
        ({"mb_per_sec"}, "MB/s"),
        ({"kb_per_sec", "kib_s"}, "KiB/s"),
        ({"gb"}, "GB"),
        ({"mb"}, "MB"),
        ({"kbytes"}, "KiB"),
        ({"nanoseconds", "nsec", "ns"}, "ns"),
        ({"microseconds", "usec", "usecs", "us"}, "us"),
        ({"milliseconds", "msec", "ms"}, "ms"),
        ({"rtt"}, "us"),
        ({"seconds", "secs", "sec", "s"}, "s"),
        ({"percent", "percentage", "pct"}, "%"),
        ({"bits_per_second", "bps", "bitrate"}, "bit/s"),
        ({"bw", "bandwidth"}, "throughput"),
        ({"kb"}, "KB"),
        ({"packets", "packet"}, "packets"),
        ({"bytes"}, "bytes"),
    )

    for patterns, candidate in rules:
        if metric_has_any_pattern(metric_text, metric_tokens, patterns):
            unit = candidate
            break

    if not unit and metric_kind in {"reliability_percent", "utilization"}:
        unit = "% / ratio"
    elif not unit and metric_kind in {
        "count",
        "network_volume",
        "reliability_count",
        "work_count",
    }:
        unit = "count"
    elif not unit and metric_kind in {"data_volume", "io_volume"}:
        unit = "bytes"
    elif not unit and metric_kind in {
        "interval_time",
        "memory",
        "throughput",
        "time",
        "variation",
    }:
        unit = UNSPECIFIED_UNIT

    return unit


def metric_description(metric_name: str, metric_kind: str, unit: str) -> str:  # noqa: C901, PLR0912, PLR0915
    metric_text = normalize_metric_text(metric_name)
    metric_tokens = set(metric_text.split("_"))
    unit_text = f" Values are reported in {unit}." if unit and unit != UNSPECIFIED_UNIT else ""

    if is_variability_metric(metric_text):
        description = "Variability or standard deviation of the measured metric across samples."
    elif metric_matches(metric_text, metric_tokens, "iops"):
        description = "Number of disk input/output operations completed per second."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"bogo_ops_per_sec_cpu"}):
        description = (
            "stress-ng throughput: bogo operations completed per second of combined "
            "user and system CPU time."
        )
    elif metric_has_any_pattern(metric_text, metric_tokens, {"bogo_ops_per_sec_real"}):
        description = (
            "stress-ng throughput: bogo operations completed per second of wall-clock time."
        )
    elif metric_has_any_pattern(metric_text, metric_tokens, {"bogo_ops_s_usr_sys_time"}):
        description = (
            "stress-ng throughput: bogo operations completed per second of combined "
            "user and system CPU time."
        )
    elif metric_has_any_pattern(metric_text, metric_tokens, {"bogo_ops_s_real_time"}):
        description = (
            "stress-ng throughput: bogo operations completed per second of wall-clock time."
        )
    elif metric_has_any_pattern(metric_text, metric_tokens, {"bogo_ops_per_sec", "bogo_ops_s"}):
        description = "stress-ng throughput: bogo operations completed per second."
    elif metric_matches(metric_text, metric_tokens, "bogo_ops"):
        description = "Number of stress-ng bogo operations completed during the run."
    elif metric_has_any_pattern(
        metric_text,
        metric_tokens,
        {"bits_per_second", "bps", "bitrate"},
    ):
        description = "Network throughput: number of bits transferred per second."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"bw", "bandwidth", "kib_s"}):
        description = "Data throughput reported by the benchmark."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"ops_per_sec", "ops_s"}):
        description = "Operations completed per second."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"records_per_second"}):
        description = "Records processed per second."
    elif metric_has_any_pattern(
        metric_text,
        metric_tokens,
        {"microseconds_per_op", "usec_per_op", "usecs_per_op"},
    ):
        description = "Average time spent per operation."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"real_time_secs"}):
        description = "Wall-clock time elapsed during the measured run."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"usr_time_secs"}):
        description = "User-mode CPU time consumed by the measured process."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"sys_time_secs"}):
        description = "Kernel-mode CPU time consumed by the measured process."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"average_sec", "avg_sec"}):
        description = "Average elapsed time for the measured operation."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"clat"}):
        description = "I/O completion latency: time from submitting an operation to completion."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"slat"}):
        description = "I/O submission latency: time spent submitting an operation to the kernel."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"rtt"}):
        description = "Round-trip latency reported by the network test."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"latency", "lat", "rtt"}):
        description = "Latency of the measured operation or round trip."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"runtime", "duration", "time"}):
        description = "Elapsed time required to complete the measured operation."
    elif is_transfer_volume_metric(metric_text, metric_tokens):
        description = "Amount of data transferred during the measured run."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"rss"}):
        description = "Resident memory footprint used by the measured process or stressor."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"disk_usage", "disk_util"}):
        description = "Disk utilization reported by the tool."
    elif metric_has_any_pattern(
        metric_text,
        metric_tokens,
        {"cpu", "cpu_used", "util", "usage"},
    ):
        description = "CPU utilization reported by the tool."
    elif metric_has_any_pattern(
        metric_text,
        metric_tokens,
        {"loss_percent", "lost_percent", "packet_loss_percent"},
    ):
        description = "Percentage of packets or items lost during the measured interval."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"lost_packets"}):
        description = "Number of network packets lost during the measured interval."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"packets"}):
        description = "Number of network packets transferred during the measured interval."
    elif (
        metric_has_any_pattern(metric_text, metric_tokens, {"bytes"}) and metric_kind == "io_volume"
    ):
        description = "Amount of disk I/O data completed during the measured run."
    elif (
        metric_has_any_pattern(
            metric_text,
            metric_tokens,
            {"bytes"},
        )
        and metric_kind == "network_volume"
    ):
        description = "Amount of network data transferred during the measured interval."
    elif metric_has_any_pattern(metric_text, metric_tokens, {"bytes"}):
        description = "Amount of data processed or transferred during the measured run."
    elif metric_kind == "interval_time":
        description = "Measurement interval duration reported by the tool."
    elif metric_kind == "reliability_percent":
        description = "Percentage of failures, lost items or other problem events."
    elif metric_kind == "reliability_count":
        description = "Number of failures, errors, lost items or other problem events."
    elif metric_kind == "memory":
        description = "Memory consumption or size footprint reported by the tool."
    elif metric_kind == "network_volume":
        description = "Amount of network data transferred during the measured interval."
    elif metric_kind == "io_volume":
        description = "Amount of disk I/O data completed during the measured run."
    elif metric_kind == "data_volume":
        description = "Amount of data processed or transferred during the measured run."
    elif metric_kind == "variation":
        description = "Variability or standard deviation of the measured metric across samples."
    elif metric_kind == "work_count":
        description = "Amount of benchmark work completed during the measured run."
    elif metric_kind == "throughput":
        description = "Work completed per unit of time."
    elif metric_kind == "time":
        description = "Elapsed time, latency or wait duration for the measured operation."
    elif metric_kind == "utilization":
        description = "Resource utilization, usage rate or ratio reported by the tool."
    elif metric_kind == "count":
        description = "Discrete count or total reported by the tool."
    else:
        description = "Numeric scalar value reported by the test."

    return f"{description}{unit_text}"


def metric_direction_for(
    metric_kind: str,
    metric_text: str,
    metric_tokens: set[str],
) -> tuple[MetricDirection, bool | None]:
    if metric_kind == "variation" or is_variability_metric(metric_text):
        result: tuple[MetricDirection, bool | None] = ("lower", False)
    elif metric_kind == "interval_time":
        result = ("context_dependent", None)
    elif metric_has_any_pattern(metric_text, metric_tokens, ISSUE_COUNT_PATTERNS):
        result = ("lower", False)
    elif metric_has_any_pattern(metric_text, metric_tokens, CONTEXT_DEPENDENT_PATTERNS):
        result = ("context_dependent", None)
    elif metric_has_any_pattern(metric_text, metric_tokens, THROUGHPUT_PATTERNS):
        result = ("higher", True)
    elif metric_has_any_pattern(metric_text, metric_tokens, TIME_PATTERNS):
        result = ("lower", False)
    elif metric_kind == "work_count":
        result = ("higher", True)
    elif metric_kind == "reliability_count":
        result = ("lower", False)
    else:
        result = ("context_dependent", None)
    return result


def build_chart_specs(
    buckets: list[MetricBucket],
    groups: list[ComparisonGroup],
    *,
    max_charts: int,
) -> list[ChartSpec]:
    candidates: list[tuple[int, str, ChartSpec]] = []

    for bucket in buckets:
        points: list[ChartPoint] = []
        for group in groups:
            values_by_distro = bucket.values.get(group.group_id, {})
            poky_stats = distribution_stats_or_none(values_by_distro.get("poky", []))
            suse_stats = distribution_stats_or_none(values_by_distro.get("suse", []))
            if poky_stats is not None and suse_stats is not None:
                points.append(
                    ChartPoint(
                        category=group.short_label,
                        poky=poky_stats.mean,
                        suse=suse_stats.mean,
                        poky_stats=poky_stats,
                        suse_stats=suse_stats,
                    ),
                )

        if not points:
            continue

        chart_kind, y_axis_title, reason = choose_chart_style(bucket, points)
        chart = ChartSpec(
            chart_no=0,
            logical_test=bucket.logical_test,
            source_sheet=bucket.source_sheet,
            metric_name=bucket.metric_name,
            chart_kind=chart_kind,
            y_axis_title=y_axis_title,
            reason=reason,
            points=tuple(points),
        )
        score = chart_score(bucket, len(points))
        sort_key = f"{bucket.logical_test}|{bucket.source_sheet}|{bucket.metric_name}"
        candidates.append((score, sort_key, chart))

    candidates.sort(key=lambda item: (-item[0], item[1]))
    if max_charts > 0:
        candidates = candidates[:max_charts]

    result: list[ChartSpec] = []
    for chart_no, (_, _, chart) in enumerate(candidates, start=1):
        result.append(
            ChartSpec(
                chart_no=chart_no,
                logical_test=chart.logical_test,
                source_sheet=chart.source_sheet,
                metric_name=chart.metric_name,
                chart_kind=chart.chart_kind,
                y_axis_title=chart.y_axis_title,
                reason=chart.reason,
                points=chart.points,
            ),
        )
    return result


def choose_chart_style(bucket: MetricBucket, points: list[ChartPoint]) -> tuple[str, str, str]:
    points_count = len(points)
    text = normalize_metric_text(
        f"{bucket.logical_test} {bucket.source_sheet} {bucket.metric_name}",
    )
    tokens = set(text.split("_"))
    y_axis_title = y_axis_title_for(text, tokens)

    if is_iperf_test_bucket(bucket):
        style = (
            "candlestick",
            y_axis_title,
            "iperf metrics are shown as candlestick charts",
        )
    elif is_stress_ng_bucket(bucket):
        style = trend_chart_style(
            points_count,
            y_axis_title,
            (
                "stress-ng metrics are compared consistently inside one test type "
                "by run-to-run trend"
            ),
        )
    elif metric_has_any_pattern(text, tokens, {"systemd", "systemd_analyze"}):
        style = (
            "histogram",
            "Mean time, sec; lower is better",
            (
                "systemd-analyze is a set of boot-time scalar values, so grouped "
                "columns compare Poky and SUSE directly"
            ),
        )
    elif metric_has_any_pattern(text, tokens, COUNT_PATTERNS):
        style = (
            "column",
            "Mean count",
            (
                "count/error metrics are discrete scalar values, so grouped columns "
                "make Poky/SUSE differences easy to read"
            ),
        )
    elif metric_has_any_pattern(text, tokens, MEMORY_PATTERNS):
        style = trend_chart_style(
            points_count,
            "Mean memory/size",
            "memory/size metrics are compared by mean values across comparable runs",
        )
    elif metric_has_any_pattern(
        text,
        tokens,
        THROUGHPUT_PATTERNS | PERCENT_PATTERNS | TIME_PATTERNS,
    ):
        style = trend_chart_style(
            points_count,
            y_axis_title,
            "this metric is compared by mean values across comparable runs",
        )
    else:
        style = trend_chart_style(
            points_count,
            y_axis_title,
            "generic scalar metric comparison by mean values",
        )
    return style


def trend_chart_style(points_count: int, y_axis_title: str, reason: str) -> tuple[str, str, str]:
    if points_count >= MIN_TREND_POINTS:
        return ("line", y_axis_title, reason)
    return ("column", y_axis_title, reason)


def y_axis_title_for(text: str, tokens: set[str]) -> str:
    if metric_has_any_pattern(text, tokens, IPERF_PATTERNS) and is_interval_time_text(
        text,
        tokens,
    ):
        title = "Mean time"
    elif metric_has_any_pattern(text, tokens, PERCENT_PATTERNS):
        title = "Mean, % / ratio"
    elif metric_has_any_pattern(text, tokens, THROUGHPUT_PATTERNS):
        title = "Mean throughput; higher is better"
    elif metric_has_any_pattern(text, tokens, TIME_PATTERNS):
        title = "Mean time; lower is better"
    elif metric_has_any_pattern(text, tokens, MEMORY_PATTERNS):
        title = "Mean memory/size"
    elif metric_has_any_pattern(text, tokens, COUNT_PATTERNS):
        title = "Mean count"
    else:
        title = "Mean"
    return title


def chart_score(bucket: MetricBucket, points_count: int) -> int:
    text = normalize_metric_text(
        f"{bucket.logical_test} {bucket.source_sheet} {bucket.metric_name}",
    )
    tokens = set(text.split("_"))
    score = 0
    for pattern, pattern_score in CHART_SCORE_RULES:
        if pattern in IPERF_PATTERNS and not is_iperf_test_bucket(bucket):
            continue
        if metric_matches(text, tokens, pattern):
            score = max(score, pattern_score)
    if points_count >= MIN_TREND_POINTS:
        score += 10
    return score


def is_iperf_test_bucket(bucket: MetricBucket) -> bool:
    test_text = normalize_metric_text(f"{bucket.logical_test} {bucket.source_sheet}")
    test_tokens = set(test_text.split("_"))
    return metric_has_any_pattern(test_text, test_tokens, IPERF_PATTERNS)


def is_iperf_metric(
    bucket: MetricBucket,
    metric_text: str,
    metric_tokens: set[str],
) -> bool:
    return is_iperf_test_bucket(bucket) or metric_has_any_pattern(
        metric_text,
        metric_tokens,
        IPERF_PATTERNS,
    )


def is_fio_test_bucket(bucket: MetricBucket) -> bool:
    test_text = normalize_metric_text(f"{bucket.logical_test} {bucket.source_sheet}")
    test_tokens = set(test_text.split("_"))
    return metric_has_any_pattern(test_text, test_tokens, FIO_PATTERNS)


def is_stress_ng_bucket(bucket: MetricBucket) -> bool:
    test_text = normalize_metric_text(f"{bucket.logical_test} {bucket.source_sheet}")
    test_tokens = set(test_text.split("_"))
    return metric_has_any_pattern(test_text, test_tokens, STRESS_NG_PATTERNS)


def write_comparison_selection(
    workbook: Workbook,
    used_sheet_names: set[str],
    groups: list[ComparisonGroup],
    chart_count: int,
) -> None:
    ws = workbook.create_sheet(unique_sheet_name(COMPARISON_SELECTION_SHEET, used_sheet_names))
    rows: list[list[Any]] = [
        [
            "group_order",
            "group_id",
            "group_label",
            "short_label",
            "experiment_order",
            "experiment_id",
            "distro",
            "configuration_id",
            "description",
            "type",
            "started_at",
        ],
    ]
    for group in groups:
        for experiment_order, distro in enumerate(DISTROS, start=1):
            info = group.experiments[distro]
            rows.append(
                [
                    group.order,
                    group.group_id,
                    group.label,
                    group.short_label,
                    experiment_order,
                    info.experiment_id,
                    info.distro,
                    info.configuration_id,
                    info.description,
                    info.experiment_type,
                    info.started_at,
                ],
            )
    rows.append([])
    rows.append(["charts_built", chart_count] + [None] * 9)
    write_styled_table(ws, rows)


def write_comparison_metrics(
    workbook: Workbook,
    used_sheet_names: set[str],
    metric_definitions: list[MetricDefinition],
) -> None:
    ws = workbook.create_sheet(unique_sheet_name(COMPARISON_METRICS_SHEET, used_sheet_names))
    rows: list[list[Any]] = [
        [
            "metric_name",
            "base_metric_name",
            "metric_kind",
            "unit",
            "description",
            "preferred_direction",
            "higher_is_better",
        ],
    ]
    rows.extend(
        [
            definition.metric_name,
            definition.base_metric_name,
            definition.metric_kind,
            definition.unit,
            definition.description,
            definition.preferred_direction,
            definition.higher_is_better,
        ]
        for definition in metric_definitions
    )
    write_styled_table(ws, rows)
    style_comparison_metrics_sheet(ws, len(rows))


def style_comparison_metrics_sheet(ws: Worksheet, row_count: int) -> None:
    """Make the metrics dictionary sheet readable in Excel.

    The shared table style keeps columns compact, while this sheet contains long
    metric descriptions. Wider semantic columns and wrapped description cells keep
    the generated dictionary scannable without hiding text behind Excel defaults.
    """
    widths = {
        "A": 34,
        "B": 34,
        "C": 18,
        "D": 14,
        "E": 88,
        "F": 22,
        "G": 18,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    if row_count < FIRST_DATA_ROW:
        return

    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, max_row=row_count, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_comparison_index(
    workbook: Workbook,
    used_sheet_names: set[str],
    charts: list[ChartSpec],
) -> None:
    ws = workbook.create_sheet(unique_sheet_name(COMPARISON_INDEX_SHEET, used_sheet_names))
    rows: list[list[Any]] = [
        [
            "chart_no",
            "title",
            "test_type",
            "test_name",
            "metric_name",
            "points",
        ],
    ]
    rows.extend(
        [
            chart.chart_no,
            chart_title(chart),
            chart.source_sheet,
            chart.logical_test,
            chart.metric_name,
            len(chart.points),
        ]
        for chart in charts
    )
    write_styled_table(ws, rows)


def write_comparison_data(
    workbook: Workbook,
    used_sheet_names: set[str],
    charts: list[ChartSpec],
) -> None:
    ws = workbook.create_sheet(unique_sheet_name(COMPARISON_DATA_SHEET, used_sheet_names))
    rows: list[list[Any]] = [
        [
            "chart_no",
            "logical_test",
            "source_sheet",
            "metric_name",
            "category",
            "poky",
            "suse",
        ],
    ]
    rows.extend(
        [
            chart.chart_no,
            chart.logical_test,
            chart.source_sheet,
            chart.metric_name,
            point.category,
            point.poky,
            point.suse,
        ]
        for chart in charts
        for point in chart.points
    )
    write_styled_table(ws, rows)


def write_comparison_charts(
    workbook: Workbook,
    used_sheet_names: set[str],
    charts: list[ChartSpec],
    *,
    charts_per_sheet: int,
) -> None:
    if not charts:
        return

    charts_per_sheet = max(1, charts_per_sheet)
    charts_by_test_type = group_charts_by_test_type(charts)
    for test_type, test_type_charts in charts_by_test_type.items():
        for start in range(0, len(test_type_charts), charts_per_sheet):
            sheet_number = start // charts_per_sheet + 1
            sheet_name = chart_sheet_name(test_type, sheet_number)
            ws = workbook.create_sheet(unique_sheet_name(sheet_name, used_sheet_names))
            setup_chart_sheet(ws)
            current_row = 1
            for chart in test_type_charts[start : start + charts_per_sheet]:
                write_chart_block(ws, current_row, chart)
                current_row += max(chart_block_height(chart), 24)


def group_charts_by_test_type(charts: list[ChartSpec]) -> dict[str, list[ChartSpec]]:
    grouped: dict[str, list[ChartSpec]] = {}
    for chart in charts:
        test_type = chart.source_sheet or "test"
        grouped.setdefault(test_type, []).append(chart)
    return grouped


def chart_sheet_name(test_type: str, sheet_number: int) -> str:
    base_name = f"{CHART_SHEET_PREFIX}_{test_type or 'test'}"
    suffix = f"_{sheet_number}" if sheet_number > 1 else ""
    cleaned_base = sanitize_sheet_name(base_name)
    if suffix:
        return f"{cleaned_base[: MAX_SHEET_NAME_LENGTH - len(suffix)]}{suffix}"
    return cleaned_base


def chart_block_height(chart: ChartSpec) -> int:
    row_count = (
        len(candlestick_rows(chart)) if chart.chart_kind == "candlestick" else len(chart.points)
    )
    return max(row_count + 8, 24)


def setup_chart_sheet(ws: Worksheet) -> None:
    widths = {
        "A": 24,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 18,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def write_chart_block(ws: Worksheet, start_row: int, chart: ChartSpec) -> None:
    if chart.chart_kind == "candlestick":
        write_candlestick_chart_block(ws, start_row, chart)
        return

    title = chart_title(chart)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)

    header_row = start_row + 2
    headers = ["run", "poky", "suse"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    data_start = header_row + 1
    for offset, point in enumerate(chart.points):
        row = data_start + offset
        ws.cell(row=row, column=1, value=point.category)
        ws.cell(row=row, column=2, value=point.poky)
        ws.cell(row=row, column=3, value=point.suse)

    data_end = data_start + len(chart.points) - 1
    excel_chart = build_excel_chart(chart)
    excel_chart.add_data(
        Reference(ws, min_col=2, min_row=header_row, max_col=3, max_row=data_end),
        titles_from_data=True,
    )
    excel_chart.set_categories(Reference(ws, min_col=1, min_row=data_start, max_row=data_end))
    ws.add_chart(excel_chart, f"E{start_row}")


def write_candlestick_chart_block(ws: Worksheet, start_row: int, chart: ChartSpec) -> None:
    title = chart_title(chart)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)

    header_row = start_row + 2
    headers = ["run_distro", "open_q1", "high_max", "low_min", "close_q3", "mean", "samples"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    rows = candlestick_rows(chart)
    data_start = header_row + 1
    for offset, row_values in enumerate(rows):
        row = data_start + offset
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=row, column=col, value=value)

    data_end = data_start + len(rows) - 1
    excel_chart = build_excel_chart(chart)
    excel_chart.add_data(
        Reference(ws, min_col=2, min_row=header_row, max_col=5, max_row=data_end),
        titles_from_data=True,
    )
    excel_chart.set_categories(Reference(ws, min_col=1, min_row=data_start, max_row=data_end))
    style_candlestick_excel_chart(excel_chart)
    ws.add_chart(excel_chart, f"I{start_row}")


def candlestick_rows(chart: ChartSpec) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for point in chart.points:
        if point.poky_stats is not None:
            rows.append(candlestick_row(f"{point.category} poky", point.poky_stats))
        if point.suse_stats is not None:
            rows.append(candlestick_row(f"{point.category} suse", point.suse_stats))
    return rows


def candlestick_row(label: str, stats: DistributionStats) -> list[Any]:
    return [label, stats.q1, stats.high, stats.low, stats.q3, stats.mean, stats.samples]


def build_excel_chart(chart: ChartSpec) -> BarChart | LineChart | AreaChart | StockChart:
    if chart.chart_kind == "line":
        excel_chart = LineChart()
    elif chart.chart_kind == "area":
        excel_chart = AreaChart()
    elif chart.chart_kind == "candlestick":
        excel_chart = StockChart()
        excel_chart.hiLowLines = ChartLines()
        excel_chart.upDownBars = UpDownBars()
    else:
        excel_chart = BarChart()
        excel_chart.type = "bar" if chart.chart_kind == "bar" else "col"
        excel_chart.grouping = "clustered"
        excel_chart.overlap = 0

    excel_chart.title = chart_title(chart)[:255]
    excel_chart.y_axis.title = chart.y_axis_title
    excel_chart.x_axis.title = "Run"
    excel_chart.height = 8
    excel_chart.width = 18
    if should_show_legend(chart):
        excel_chart.legend.position = "r"
    else:
        excel_chart.legend = None
    return excel_chart


def should_show_legend(chart: ChartSpec) -> bool:
    return chart.chart_kind != "candlestick"


def style_candlestick_excel_chart(excel_chart: StockChart) -> None:
    for series in excel_chart.series:
        series.graphicalProperties.line.noFill = True
        series.graphicalProperties.noFill = True
        if series.marker is not None:
            series.marker.symbol = "none"
            series.marker.graphicalProperties.line.noFill = True
            series.marker.graphicalProperties.noFill = True


def chart_title(chart: ChartSpec) -> str:
    test_type = str(chart.source_sheet or "test").strip()
    test_name = str(chart.logical_test or "test").strip()
    metric_name = str(chart.metric_name or "metric").strip()

    if normalize_metric_text(test_type) == normalize_metric_text(test_name):
        return f"{test_type}: {metric_name}"

    return f"{test_type} ({test_name}): {metric_name}"


def write_table_sheet(ws: Worksheet, sheet: SheetRows) -> None:
    rows = [sheet.headers, *sheet.rows]
    write_styled_table(ws, rows)


def write_styled_table(ws: Worksheet, rows: list[list[Any]]) -> None:
    write_matrix(ws, rows)
    if not rows or not rows[0]:
        return
    style_table(ws, len(rows), len(rows[0]))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def distribution_stats_or_none(values: list[float]) -> DistributionStats | None:
    prepared = sorted(value for value in values if is_nonzero_number(value))
    if not prepared:
        return None
    return DistributionStats(
        mean=statistics.fmean(prepared),
        low=prepared[0],
        q1=percentile(prepared, 0.25),
        q3=percentile(prepared, 0.75),
        high=prepared[-1],
        samples=len(prepared),
    )


def percentile(sorted_values: list[float], fraction: float) -> float:
    if not sorted_values:
        msg = "Cannot calculate percentile for an empty list."
        raise ValueError(msg)
    if len(sorted_values) == 1:
        return sorted_values[0]

    position = (len(sorted_values) - 1) * fraction
    lower_index = math.floor(position)
    upper_index = math.ceil(position)
    if lower_index == upper_index:
        return sorted_values[lower_index]

    lower_value = sorted_values[lower_index]
    upper_value = sorted_values[upper_index]
    weight = position - lower_index
    return lower_value + (upper_value - lower_value) * weight
