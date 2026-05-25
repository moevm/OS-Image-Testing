from __future__ import annotations

import logging
import re
import statistics
from dataclasses import dataclass
from typing import TYPE_CHECKING, Any, Final

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from imgtests.reporting.distro_comparison_common import (
    COUNT_PATTERNS,
    IPERF_PATTERNS,
    LABEL_COLUMNS,
    MEMORY_PATTERNS,
    NONZERO_EPSILON,
    PERCENT_PATTERNS,
    STRESS_NG_PATTERNS,
    THROUGHPUT_PATTERNS,
    TIME_PATTERNS,
    ComparisonGroup,
    MetricBucket,
    MetricExtractionOptions,
    build_comparison_groups,
    build_configuration_distro_map,
    build_experiment_info_map,
    metric_has_any_pattern,
    normalize_metric_text,
    style_table,
    unique_sheet_name,
    write_matrix,
)
from imgtests.reporting.distro_comparison_common import (
    extract_metric_buckets as extract_common_metric_buckets,
)
from imgtests.reporting.distro_comparison_common import (
    resolve_output_path as resolve_common_output_path,
)

if TYPE_CHECKING:
    from collections.abc import Sequence
    from pathlib import Path

logger = logging.getLogger(__name__)

BASELINE_DISTRO: Final = "suse"
COMPARED_DISTRO: Final = "poky"
OUTPUT_NAME: Final = "comparison_status"
SUMMARY_SHEET: Final = "summary"
TEST_STATUS_SHEET: Final = "test_status"
METRIC_STATUS_SHEET: Final = "metric_status"

DEFAULT_EPSILON_PERCENT: Final = 1.0

HIGHER_IS_BETTER: Final = "higher"
LOWER_IS_BETTER: Final = "lower"
NOT_COMPARABLE_DIRECTION: Final = "not_comparable"

STATUS_OK: Final = "PASS"
STATUS_FAIL: Final = "FAIL"
STATUS_NOT_COMPARABLE: Final = "NOT_COMPARABLE"

CHANGE_IMPROVED: Final = "improved"
CHANGE_DEGRADED: Final = "degraded"
CHANGE_UNCHANGED: Final = "unchanged"
CHANGE_NOT_COMPARABLE: Final = "not_comparable"

RAW_INCREASED: Final = "increased"
RAW_DECREASED: Final = "decreased"
RAW_UNCHANGED: Final = "unchanged"
RAW_UNAVAILABLE: Final = "unavailable"

FLATTENED_DUPLICATE_SUFFIX_RE: Final = re.compile(r"_\d+$")
PTS_TEST_RUN_TIMES_METRIC: Final = "results_test_run_times"
PTS_TEST_RUN_TIMES_RE: Final = re.compile(r"^results_test_run_times(?:_\d+)?$")

IPERF_INTERVAL_ANY_RE: Final = re.compile(r"(^|_)intervals(_|$)")

IPERF_PREFIX_RE: Final = re.compile(r"^iperf3_")
IPERF_TARGET_METRIC_RE: Final = re.compile(
    r"^("
    r"(client|server)_(sum|sent|received|streams_sender|streams_receiver|streams_udp)_"
    r"(bps|bytes|packets|retransmits|rtt|mean_rtt|min_rtt|max_rtt|jitter_ms|"
    r"lost_packets|lost_percent|out_of_order)"
    r")$",
)
IPERF_LOWER_IS_BETTER_RE: Final = re.compile(
    r"(^|_)(jitter_ms|lost_packets|lost_percent|out_of_order|"
    r"retransmits|rtt|mean_rtt|min_rtt|max_rtt)$",
)
IPERF_CPU_HOST_METRIC_RE: Final = re.compile(
    r"^(iperf3_)?(client|server)_cpu_host_(total|user|system)$",
)
IPERF_TECHNICAL_METRIC_PARTS: Final = {
    "pmtu",
    "rcvbuf_actual",
    "seconds",
    "snd_cwnd",
    "snd_wnd",
    "sndbuf_actual",
    "sock_bufsize",
    "target_bitrate",
    "test_bidir",
    "test_block_size",
    "test_blocks",
    "test_bytes",
    "test_duration_sec",
    "test_interval",
    "test_omit",
    "test_reverse",
    "test_streams",
    "test_target_bitrate",
}

NON_PERFORMANCE_SOURCE_SHEETS: Final = {
    "planned_stage",
    "task_result_for_stage_warmup",
    "unnamed_util_run_result",
}

STRESS_NG_TARGET_METRICS: Final = {
    "bogo_ops",
    "bogo_ops_s_real_time",
    "bogo_ops_s_usr_sys_time",
    "stress_ng_metrics_bogo_ops",
    "stress_ng_metrics_bogo_ops_s_real_time",
    "stress_ng_metrics_bogo_ops_s_usr_sys_time",
}

PERF_TARGET_METRICS: Final = {
    "gb_per_sec_default",
    "gb_per_sec_movsq_based",
    "gb_per_sec_unrolled",
    "ops_per_sec",
    "usecs_per_op",
    "perf_metrics_gb_per_sec_default",
    "perf_metrics_gb_per_sec_movsq_based",
    "perf_metrics_gb_per_sec_unrolled",
    "perf_metrics_ops_per_sec",
    "perf_metrics_usecs_per_op",
}

SYSTEMD_TARGET_METRICS: Final = {
    "firmware_time",
    "loader_time",
    "initrd_time",
    "kernel_time",
    "userspace_time",
    "total_time",
}

FIO_TARGET_METRIC_RE: Final = re.compile(
    r"^(read|write|trim)_("
    r"bw|bw_bytes|iops|io_bytes|io_kbytes|total_ios|"
    r"(slat|clat|lat)_(ns|us)_(min|max|mean|stddev)"
    r")$",
)
FIO_LOWER_IS_BETTER_RE: Final = re.compile(
    r"^(read|write|trim)_(slat|clat|lat)_(ns|us)_(min|max|mean|stddev)$",
)

LOWER_IS_BETTER_TARGETS: Final = {
    "firmware_time",
    "loader_time",
    "initrd_time",
    "kernel_time",
    "userspace_time",
    "total_time",
    "usecs_per_op",
    "perf_metrics_usecs_per_op",
}

LOWER_IS_BETTER_PATTERNS: Final = {
    "broken",
    "clat",
    "drop",
    "dropped",
    "elapsed",
    "error",
    "errors",
    "fail",
    "failed",
    "failure",
    "failures",
    "fault",
    "faults",
    "jitter",
    "jitter_ms",
    "latency",
    "lost",
    "lost_packets",
    "lost_percent",
    "oom",
    "out_of_order",
    "context_switches",
    "retransmit",
    "retransmits",
    "retry",
    "retries",
    "rtt",
    "skip",
    "skipped",
    "slat",
    "timeout",
    "timeouts",
    "untrustworthy",
}

HIGHER_IS_BETTER_PATTERNS: Final = {
    "bandwidth",
    "bitrate",
    "bits_per_second",
    "bogo_ops",
    "bogo_ops_s",
    "bps",
    "gb_per_sec",
    "iops",
    "io_bytes",
    "io_kbytes",
    "ops_per_sec",
    "ops_s",
    "packets",
    "passed",
    "passrate",
    "pps",
    "requests",
    "requests_per_second",
    "success",
    "successes",
    "total_ios",
}

RESOURCE_USAGE_PATTERNS: Final = {
    "cpu",
    "cpu_used",
    "host_system",
    "host_total",
    "host_user",
    "remote_system",
    "remote_total",
    "remote_user",
    "rss",
    "usage",
    "util",
}

TECHNICAL_OR_ID_PATTERNS: Final = {
    "addr",
    "address",
    "code",
    "cookie",
    "fd",
    "gid",
    "host",
    "hostname",
    "id",
    "ip",
    "local",
    "mss",
    "pid",
    "port",
    "rc",
    "remote",
    "return",
    "returncode",
    "retval",
    "socket",
    "status",
    "timestamp",
    "tos",
    "uid",
    "unix",
    "version",
}


@dataclass(frozen=True)
class DistroComparisonStatusOptions:
    output_path: Path | None = None
    experiment_ids: Sequence[str] | None = None
    latest_pair_only: bool = False
    epsilon_percent: float = DEFAULT_EPSILON_PERCENT


@dataclass(frozen=True)
class MetricDirection:
    better_direction: str
    interpretation: str
    reason: str
    comparable: bool


@dataclass(frozen=True)
class MeanStats:
    mean: float
    sample_count: int


@dataclass(frozen=True)
class MetricStatusRow:
    status: str
    performance_change: str
    raw_value_change: str
    group_order: int
    group_label: str
    source_sheet: str
    test_name: str
    metric_name: str
    baseline_distro: str
    baseline_mean: float
    baseline_sample_count: int
    compared_distro: str
    compared_mean: float
    compared_sample_count: int
    better_direction: str
    epsilon_percent: float
    value_delta_percent: float | None
    performance_delta_percent: float | None


@dataclass(frozen=True)
class TestStatusRow:
    status: str
    group_order: int
    group_label: str
    source_sheet: str
    test_name: str
    baseline_distro: str
    compared_distro: str
    epsilon_percent: float
    metrics_total: int
    passed_metrics: int
    failed_metrics: int
    mean_value_delta_percent: float | None
    mean_performance_delta_percent: float | None
    failed_metric_names: str


def export_distro_comparison_status_to_excel(
    input_path: Path,
    options: DistroComparisonStatusOptions | None = None,
) -> Path:
    options = options or DistroComparisonStatusOptions()
    source_wb = load_workbook(input_path, read_only=True, data_only=True)
    config_distro = build_configuration_distro_map(source_wb)
    experiment_info = build_experiment_info_map(source_wb, config_distro)
    groups = build_comparison_groups(
        experiment_info,
        experiment_ids=list(options.experiment_ids) if options.experiment_ids is not None else None,
        latest_pair_only=options.latest_pair_only,
    )
    logger.info("Comparable Poky/SUSE groups: %s", len(groups))

    rows: list[MetricStatusRow] = []
    if groups:
        buckets = extract_status_metric_buckets(source_wb, config_distro, groups)
        logger.info("Metric buckets with data: %s", len(buckets))
        rows = build_metric_status_rows(
            buckets,
            groups,
            epsilon_percent=options.epsilon_percent,
        )
        logger.info("Metric status rows: %s", len(rows))

    result_path = resolve_common_output_path(input_path, options.output_path, OUTPUT_NAME)
    write_status_workbook(
        result_path,
        rows,
        groups,
        epsilon_percent=max(0.0, options.epsilon_percent),
    )
    return result_path


def extract_status_metric_buckets(
    workbook: Any,
    config_distro: dict[str, str],
    groups: list[ComparisonGroup],
) -> list[MetricBucket]:
    return extract_common_metric_buckets(
        workbook,
        config_distro,
        groups,
        MetricExtractionOptions(
            column_selector=status_metric_columns,
            name_mapper=status_metric_bucket_name,
            predicate=should_collect_metric_status,
            require_nonzero=False,
        ),
    )


def status_metric_columns(headers: list[str]) -> list[int]:
    metric_indexes: list[int] = []
    for index, header in enumerate(headers):
        if not header or header in LABEL_COLUMNS:
            continue
        if normalize_metric_text(header):
            metric_indexes.append(index)
    return metric_indexes


def should_collect_metric_status(
    source_sheet: str,
    logical_test: str,
    metric_name: str,
) -> bool:
    return target_metric_kind_for(source_sheet, logical_test, metric_name) is not None


def build_metric_status_rows(
    buckets: list[MetricBucket],
    groups: list[ComparisonGroup],
    *,
    epsilon_percent: float,
) -> list[MetricStatusRow]:
    normalized_epsilon = max(0.0, epsilon_percent)
    rows: list[MetricStatusRow] = []
    sorted_buckets = sorted(
        buckets,
        key=lambda bucket: (bucket.source_sheet, bucket.logical_test, bucket.metric_name),
    )

    for bucket in sorted_buckets:
        if not should_write_metric_status(bucket):
            continue
        direction = metric_direction(bucket)
        if not direction.comparable:
            continue
        for group in groups:
            values_by_distro = bucket.values.get(group.group_id, {})
            baseline_stats = mean_stats_or_none(values_by_distro.get(BASELINE_DISTRO, []))
            compared_stats = mean_stats_or_none(values_by_distro.get(COMPARED_DISTRO, []))
            if baseline_stats is None or compared_stats is None:
                continue

            value_delta = percent_delta(compared_stats.mean, baseline_stats.mean)
            performance_delta = directional_percent_delta(
                value_delta,
                direction.better_direction,
            )
            performance_change = performance_change_for_values(
                baseline_stats.mean,
                compared_stats.mean,
                direction,
                normalized_epsilon,
            )
            rows.append(
                MetricStatusRow(
                    status=status_for(performance_change),
                    performance_change=performance_change,
                    raw_value_change=raw_value_change_for(
                        value_delta,
                        normalized_epsilon,
                        baseline=baseline_stats.mean,
                        compared=compared_stats.mean,
                    ),
                    group_order=group.order,
                    group_label=group.short_label,
                    source_sheet=bucket.source_sheet,
                    test_name=bucket.logical_test,
                    metric_name=bucket.metric_name,
                    baseline_distro=BASELINE_DISTRO,
                    baseline_mean=baseline_stats.mean,
                    baseline_sample_count=baseline_stats.sample_count,
                    compared_distro=COMPARED_DISTRO,
                    compared_mean=compared_stats.mean,
                    compared_sample_count=compared_stats.sample_count,
                    better_direction=direction.better_direction,
                    epsilon_percent=normalized_epsilon,
                    value_delta_percent=value_delta,
                    performance_delta_percent=performance_delta,
                ),
            )

    return rows


def should_write_metric_status(bucket: MetricBucket) -> bool:
    return target_metric_kind(bucket) is not None


def target_metric_kind(bucket: MetricBucket) -> str | None:
    return target_metric_kind_for(bucket.source_sheet, bucket.logical_test, bucket.metric_name)


def target_metric_kind_for(source_sheet: str, logical_test: str, metric_name: str) -> str | None:
    metric_text = normalize_metric_text(metric_name)
    source_text = normalize_metric_text(source_sheet)
    test_text = normalize_metric_text(logical_test)

    if is_garbage_metric(source_text, test_text, metric_text):
        return None

    direction = inferred_metric_direction(source_text, test_text, metric_text)
    if not direction.comparable:
        return None

    return metric_kind_for(source_text, test_text, metric_text)


def is_garbage_metric(source_text: str, _test_text: str, metric_text: str) -> bool:
    if source_text in NON_PERFORMANCE_SOURCE_SHEETS:
        return True
    if is_flattened_duplicate_metric(metric_text) and not is_pts_test_run_times_metric(
        source_text,
        metric_text,
    ):
        return True
    if IPERF_INTERVAL_ANY_RE.search(metric_text):
        return True
    if has_technical_iperf_metric_part(normalized_iperf_metric_name(metric_text)):
        return True
    if is_iperf_cpu_host_metric(metric_text):
        return False
    return bool(technical_metric_direction(set(metric_text.split("_"))))


def metric_kind_for(source_text: str, test_text: str, metric_text: str) -> str:
    if is_iperf_metric_context_text(source_text, test_text, metric_text):
        return "iperf3"
    if is_perf_metric_context_text(source_text, test_text, metric_text):
        return "perf"
    if is_stress_ng_metric_context_text(source_text, test_text, metric_text):
        return "stress_ng"
    if is_fio_metric_context_text(source_text, test_text):
        return "fio"
    if source_text:
        return source_text
    return "inferred"


def inferred_metric_direction(
    source_text: str,
    test_text: str,
    metric_text: str,
) -> MetricDirection:
    text = normalize_metric_text(f"{source_text} {test_text} {metric_text}")
    tokens = set(text.split("_"))

    comparable = context_specific_metric_direction(source_text, test_text, metric_text)
    if comparable is not None:
        return comparable

    comparable = comparable_explicit_direction(text, tokens)
    if comparable is not None:
        return comparable

    comparable = generic_context_direction(text, tokens)
    if comparable is not None:
        return comparable

    non_comparable = non_comparable_explicit_direction(text, tokens)
    if non_comparable is not None:
        return non_comparable

    return MetricDirection(
        better_direction=NOT_COMPARABLE_DIRECTION,
        interpretation="unknown metric direction",
        reason="the metric is numeric but higher/lower cannot be inferred safely",
        comparable=False,
    )


def context_specific_metric_direction(  # noqa: PLR0911
    source_text: str,
    test_text: str,
    metric_text: str,
) -> MetricDirection | None:
    normalized_metric = normalized_target_metric_name(metric_text)
    metric_tokens = set(normalized_metric.split("_"))

    if is_iperf_metric_context_text(source_text, test_text, metric_text):
        if IPERF_LOWER_IS_BETTER_RE.search(normalized_metric):
            return comparable_direction(
                LOWER_IS_BETTER,
                "iperf quality/error metric",
                "lower values mean less loss, jitter, retransmits, RTT, or packet disorder",
            )
        if is_iperf_cpu_host_metric(metric_text):
            return comparable_direction(
                LOWER_IS_BETTER,
                "iperf host CPU usage metric",
                "lower values mean the same network work used less host CPU",
            )
        if metric_has_any_pattern(normalized_metric, metric_tokens, {"bps", "bytes", "packets"}):
            return comparable_direction(
                HIGHER_IS_BETTER,
                "iperf throughput metric",
                "higher values mean more network throughput or transferred data",
            )

    if is_fio_metric_context_text(source_text, test_text):
        if FIO_LOWER_IS_BETTER_RE.fullmatch(normalized_metric):
            return comparable_direction(
                LOWER_IS_BETTER,
                "fio latency metric",
                "lower values mean less IO latency",
            )
        if metric_has_any_pattern(
            normalized_metric,
            metric_tokens,
            {"bw", "io_bytes", "io_kbytes", "iops", "total_ios"},
        ):
            return comparable_direction(
                HIGHER_IS_BETTER,
                "fio throughput metric",
                "higher values mean more IO completed",
            )

    if is_perf_metric_context_text(source_text, test_text, metric_text):
        if normalized_metric in {"usecs_per_op", "perf_metrics_usecs_per_op"}:
            return comparable_direction(
                LOWER_IS_BETTER,
                "perf latency metric",
                "lower values mean each operation finished sooner",
            )
        if normalized_metric in PERF_TARGET_METRICS:
            return comparable_direction(
                HIGHER_IS_BETTER,
                "perf throughput metric",
                "higher values mean more benchmark work completed",
            )

    return None


def is_iperf_metric_context_text(source_text: str, test_text: str, metric_text: str) -> bool:
    return (
        "iperf3" in source_text
        or "iperf" in source_text
        or "iperf3" in test_text
        or "iperf" in test_text
        or metric_text.startswith("iperf3_")
    )


def is_stress_ng_metric_context_text(source_text: str, test_text: str, metric_text: str) -> bool:
    return (
        source_text.startswith("stress_ng")
        or "stress_ng" in test_text
        or metric_text.startswith("stress_ng_metrics_")
    )


def is_perf_metric_context_text(source_text: str, test_text: str, metric_text: str) -> bool:
    return source_text == "perf" or "perf" in test_text or metric_text.startswith("perf_metrics_")


def is_fio_metric_context_text(source_text: str, test_text: str) -> bool:
    return source_text.startswith("fio") or "fio" in test_text


def status_metric_bucket_name(source_sheet: str, logical_test: str, metric_name: str) -> str:
    metric_text = normalize_metric_text(metric_name)
    source_text = normalize_metric_text(source_sheet)
    if is_pts_test_run_times_metric(source_text, metric_text):
        return PTS_TEST_RUN_TIMES_METRIC
    base_metric = flattened_duplicate_base_metric(metric_text)
    if (
        base_metric is not None
        and target_metric_kind_for(source_sheet, logical_test, base_metric) is not None
    ):
        return base_metric
    return metric_name


def is_pts_test_run_times_metric(source_text: str, metric_text: str) -> bool:
    return source_text == "pts" and bool(PTS_TEST_RUN_TIMES_RE.fullmatch(metric_text))


def flattened_duplicate_base_metric(metric_text: str) -> str | None:
    if not is_flattened_duplicate_metric(metric_text):
        return None
    return FLATTENED_DUPLICATE_SUFFIX_RE.sub("", metric_text)


def is_flattened_duplicate_metric(metric_text: str) -> bool:
    return bool(FLATTENED_DUPLICATE_SUFFIX_RE.search(metric_text))


def has_technical_iperf_metric_part(metric_text: str) -> bool:
    return any(part in metric_text for part in IPERF_TECHNICAL_METRIC_PARTS)


def is_iperf_cpu_host_metric(metric_text: str) -> bool:
    return bool(IPERF_CPU_HOST_METRIC_RE.fullmatch(normalize_metric_text(metric_text)))


def normalized_iperf_metric_name(metric_text: str) -> str:
    return IPERF_PREFIX_RE.sub("", metric_text)


def normalized_target_metric_name(metric_text: str) -> str:
    metric_text = normalize_metric_text(metric_text)
    return normalized_iperf_metric_name(metric_text)


def is_target_iperf_metric(metric_text: str) -> bool:
    return bool(IPERF_TARGET_METRIC_RE.fullmatch(normalized_iperf_metric_name(metric_text)))


def is_iperf_metric_context(bucket: MetricBucket) -> bool:
    return is_iperf_metric_context_text(
        normalize_metric_text(bucket.source_sheet),
        normalize_metric_text(bucket.logical_test),
        normalize_metric_text(bucket.metric_name),
    )


def build_test_status_rows(rows: list[MetricStatusRow]) -> list[TestStatusRow]:
    grouped: dict[tuple[int, str, str, str], list[MetricStatusRow]] = {}
    for row in rows:
        key = (row.group_order, row.group_label, row.source_sheet, row.test_name)
        grouped.setdefault(key, []).append(row)

    result: list[TestStatusRow] = []
    for key in sorted(grouped):
        group_rows = grouped[key]
        failed_rows = [row for row in group_rows if row.status == STATUS_FAIL]
        value_deltas = [
            row.value_delta_percent for row in group_rows if row.value_delta_percent is not None
        ]
        performance_deltas = [
            row.performance_delta_percent
            for row in group_rows
            if row.performance_delta_percent is not None
        ]
        result.append(
            TestStatusRow(
                status=STATUS_FAIL if failed_rows else STATUS_OK,
                group_order=key[0],
                group_label=key[1],
                source_sheet=key[2],
                test_name=key[3],
                baseline_distro=BASELINE_DISTRO,
                compared_distro=COMPARED_DISTRO,
                epsilon_percent=group_rows[0].epsilon_percent,
                metrics_total=len(group_rows),
                passed_metrics=sum(1 for row in group_rows if row.status == STATUS_OK),
                failed_metrics=len(failed_rows),
                mean_value_delta_percent=(statistics.fmean(value_deltas) if value_deltas else None),
                mean_performance_delta_percent=(
                    statistics.fmean(performance_deltas) if performance_deltas else None
                ),
                failed_metric_names=", ".join(row.metric_name for row in failed_rows),
            ),
        )
    return result


def mean_stats_or_none(values: list[float]) -> MeanStats | None:
    if not values:
        return None
    return MeanStats(
        mean=statistics.fmean(values),
        sample_count=len(values),
    )


def metric_direction(bucket: MetricBucket) -> MetricDirection:
    metric_text = normalize_metric_text(bucket.metric_name)
    target_kind = target_metric_kind(bucket)

    if target_kind is None:
        return MetricDirection(
            better_direction=NOT_COMPARABLE_DIRECTION,
            interpretation="not a target metric",
            reason="the metric is not in the DB-recorded target metric allowlist",
            comparable=False,
        )

    direction = inferred_metric_direction(
        normalize_metric_text(bucket.source_sheet),
        normalize_metric_text(bucket.logical_test),
        metric_text,
    )
    return MetricDirection(
        better_direction=direction.better_direction,
        interpretation=f"{target_kind}: {direction.interpretation}",
        reason=direction.reason,
        comparable=direction.comparable,
    )


def is_lower_is_better_target(metric_text: str) -> bool:
    normalized_metric = normalized_target_metric_name(metric_text)
    return (
        normalized_metric in LOWER_IS_BETTER_TARGETS
        or bool(IPERF_LOWER_IS_BETTER_RE.search(normalized_metric))
        or bool(FIO_LOWER_IS_BETTER_RE.fullmatch(normalized_metric))
    )


def flattened_duplicate_direction(metric_text: str) -> MetricDirection | None:
    if is_flattened_duplicate_metric(metric_text):
        return MetricDirection(
            better_direction=NOT_COMPARABLE_DIRECTION,
            interpretation="flattened duplicate column",
            reason=(
                "numeric suffixes such as _2, _13 or _1000 are produced by "
                "flattening arrays and are not stable target metrics"
            ),
            comparable=False,
        )
    return None


def flattened_iperf_direction(metric_text: str) -> MetricDirection | None:
    if IPERF_INTERVAL_ANY_RE.search(metric_text):
        return MetricDirection(
            better_direction=NOT_COMPARABLE_DIRECTION,
            interpretation="flattened iperf interval column",
            reason=(
                "iperf interval fields are intermediate time-series values, "
                "not final PASS/FAIL metrics"
            ),
            comparable=False,
        )
    return None


def iperf_target_direction(bucket: MetricBucket, metric_text: str) -> MetricDirection | None:
    if not is_iperf_metric_context(bucket) or not is_target_iperf_metric(metric_text):
        return None

    base_metric = normalized_iperf_metric_name(metric_text)
    tokens = set(base_metric.split("_"))
    if metric_has_any_pattern(
        base_metric,
        tokens,
        {
            "jitter",
            "jitter_ms",
            "lost",
            "lost_packets",
            "lost_percent",
            "out_of_order",
            "retransmit",
            "retransmits",
            "rtt",
        },
    ):
        return comparable_direction(
            LOWER_IS_BETTER,
            "iperf quality/error metric",
            "lower values mean less loss, jitter, retransmits, RTT, or packet disorder",
        )

    return comparable_direction(
        HIGHER_IS_BETTER,
        "iperf throughput metric",
        "higher values mean more network throughput or transferred data",
    )


def technical_metric_direction(metric_tokens: set[str]) -> MetricDirection | None:
    if metric_tokens & TECHNICAL_OR_ID_PATTERNS:
        return MetricDirection(
            better_direction=NOT_COMPARABLE_DIRECTION,
            interpretation="technical identifier or metadata",
            reason="technical identifiers are not performance metrics",
            comparable=False,
        )
    return None


def explicit_metric_direction(text: str, tokens: set[str]) -> MetricDirection | None:
    comparable = comparable_explicit_direction(text, tokens)
    if comparable is not None:
        return comparable
    return non_comparable_explicit_direction(text, tokens)


def comparable_explicit_direction(text: str, tokens: set[str]) -> MetricDirection | None:
    if metric_has_any_pattern(text, tokens, LOWER_IS_BETTER_PATTERNS):
        return comparable_direction(
            LOWER_IS_BETTER,
            "error, loss, jitter, latency, or failure metric",
            "lower values mean fewer errors, less loss, or less waiting",
        )
    if metric_has_any_pattern(text, tokens, THROUGHPUT_PATTERNS | HIGHER_IS_BETTER_PATTERNS):
        return comparable_direction(
            HIGHER_IS_BETTER,
            "throughput, operations, packets, requests, or success metric",
            "higher values mean more work completed or more successful outcomes",
        )
    if metric_has_any_pattern(text, tokens, TIME_PATTERNS):
        return comparable_direction(
            LOWER_IS_BETTER,
            "time or duration metric",
            "lower values mean the same work finished sooner",
        )
    if metric_has_any_pattern(text, tokens, RESOURCE_USAGE_PATTERNS | MEMORY_PATTERNS):
        return comparable_direction(
            LOWER_IS_BETTER,
            "resource usage metric",
            "lower values mean less CPU, memory, or system resource consumption",
        )
    return None


def non_comparable_explicit_direction(text: str, tokens: set[str]) -> MetricDirection | None:
    if metric_has_any_pattern(text, tokens, PERCENT_PATTERNS):
        return MetricDirection(
            better_direction=NOT_COMPARABLE_DIRECTION,
            interpretation="generic percentage metric",
            reason="percentage metrics need domain context before higher/lower can be judged",
            comparable=False,
        )
    if metric_has_any_pattern(text, tokens, COUNT_PATTERNS):
        return MetricDirection(
            better_direction=NOT_COMPARABLE_DIRECTION,
            interpretation="generic count metric",
            reason="count metrics can be good or bad depending on what is counted",
            comparable=False,
        )
    return None


def generic_context_direction(text: str, tokens: set[str]) -> MetricDirection | None:
    if is_known_throughput_test(text, tokens):
        return comparable_direction(
            HIGHER_IS_BETTER,
            "generic value inside throughput benchmark",
            "the surrounding test is a throughput benchmark, so higher is treated as better",
        )
    if is_known_time_test(text, tokens):
        return comparable_direction(
            LOWER_IS_BETTER,
            "generic value inside timing benchmark",
            "the surrounding test is a timing benchmark, so lower is treated as better",
        )
    return None


def comparable_direction(
    better_direction: str,
    interpretation: str,
    reason: str,
) -> MetricDirection:
    return MetricDirection(
        better_direction=better_direction,
        interpretation=interpretation,
        reason=reason,
        comparable=True,
    )


def is_known_throughput_test(text: str, tokens: set[str]) -> bool:
    return metric_has_any_pattern(
        text,
        tokens,
        IPERF_PATTERNS | STRESS_NG_PATTERNS | {"iperf3", "network", "network_loopback"},
    )


def is_known_time_test(text: str, tokens: set[str]) -> bool:
    return metric_has_any_pattern(
        text,
        tokens,
        {"clock", "ctx_clock", "systemd", "systemd_analyze"},
    )


def percent_delta(value: float, baseline: float) -> float | None:
    if abs(baseline) <= NONZERO_EPSILON:
        if abs(value) <= NONZERO_EPSILON:
            return 0.0
        return None
    return (value - baseline) / abs(baseline) * 100


def directional_percent_delta(
    value_delta_percent: float | None,
    better_direction: str,
) -> float | None:
    if value_delta_percent is None or better_direction == NOT_COMPARABLE_DIRECTION:
        return None
    if better_direction == LOWER_IS_BETTER:
        return -value_delta_percent
    return value_delta_percent


def performance_change_for_values(
    baseline: float,
    compared: float,
    direction: MetricDirection,
    epsilon_percent: float,
) -> str:
    if not direction.comparable:
        return CHANGE_NOT_COMPARABLE

    value_delta = percent_delta(compared, baseline)
    performance_delta = directional_percent_delta(value_delta, direction.better_direction)
    if performance_delta is not None:
        return performance_change_for(performance_delta, direction, epsilon_percent)

    if abs(baseline) <= NONZERO_EPSILON:
        if abs(compared) <= NONZERO_EPSILON:
            return CHANGE_UNCHANGED
        if direction.better_direction == LOWER_IS_BETTER:
            return CHANGE_DEGRADED if compared > baseline else CHANGE_IMPROVED
        return CHANGE_IMPROVED if compared > baseline else CHANGE_DEGRADED

    return CHANGE_NOT_COMPARABLE


def raw_value_change_for(  # noqa: PLR0911
    value_delta_percent: float | None,
    epsilon_percent: float,
    *,
    baseline: float | None = None,
    compared: float | None = None,
) -> str:
    if value_delta_percent is not None:
        if abs(value_delta_percent) <= epsilon_percent:
            return RAW_UNCHANGED
        if value_delta_percent > 0:
            return RAW_INCREASED
        return RAW_DECREASED
    if baseline is not None and compared is not None:
        if abs(compared - baseline) <= NONZERO_EPSILON:
            return RAW_UNCHANGED
        if compared > baseline:
            return RAW_INCREASED
        return RAW_DECREASED
    return RAW_UNAVAILABLE


def performance_change_for(
    performance_delta_percent: float | None,
    direction: MetricDirection,
    epsilon_percent: float,
) -> str:
    if not direction.comparable or performance_delta_percent is None:
        return CHANGE_NOT_COMPARABLE
    if abs(performance_delta_percent) <= epsilon_percent:
        return CHANGE_UNCHANGED
    if performance_delta_percent > 0:
        return CHANGE_IMPROVED
    return CHANGE_DEGRADED


def status_for(performance_change: str) -> str:
    if performance_change in {CHANGE_IMPROVED, CHANGE_UNCHANGED}:
        return STATUS_OK
    if performance_change == CHANGE_DEGRADED:
        return STATUS_FAIL
    return STATUS_NOT_COMPARABLE


def write_status_workbook(
    output_path: Path,
    rows: list[MetricStatusRow],
    groups: list[ComparisonGroup],
    *,
    epsilon_percent: float,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names: set[str] = set()
    test_rows = build_test_status_rows(rows)

    write_summary_sheet(
        workbook,
        used_names,
        rows,
        test_rows,
        groups,
        epsilon_percent=epsilon_percent,
    )
    write_test_status_sheet(workbook, used_names, test_rows)
    write_metric_status_sheet(workbook, used_names, rows)

    workbook.save(output_path)


def write_summary_sheet(  # noqa: PLR0913
    workbook: Workbook,
    used_names: set[str],
    rows: list[MetricStatusRow],
    test_rows: list[TestStatusRow],
    groups: list[ComparisonGroup],
    *,
    epsilon_percent: float,
) -> None:
    metric_status_counts = count_by_status(rows)
    test_status_counts = count_by_test_status(test_rows)
    matrix: list[list[Any]] = [
        ["field", "value"],
        ["baseline_distro", BASELINE_DISTRO],
        ["compared_distro", COMPARED_DISTRO],
        ["epsilon_percent", epsilon_percent],
        ["comparison_groups", len(groups)],
        ["test_rows", len(test_rows)],
        ["metric_rows", len(rows)],
        ["test_PASS", test_status_counts.get(STATUS_OK, 0)],
        ["test_FAIL", test_status_counts.get(STATUS_FAIL, 0)],
        ["metric_PASS", metric_status_counts.get(STATUS_OK, 0)],
        ["metric_FAIL", metric_status_counts.get(STATUS_FAIL, 0)],
    ]
    ws = workbook.create_sheet(unique_sheet_name(SUMMARY_SHEET, used_names))
    write_matrix(ws, matrix)
    style_table(ws, len(matrix), len(matrix[0]))


def count_by_status(rows: list[MetricStatusRow]) -> dict[str, int]:
    result: dict[str, int] = {}
    for row in rows:
        result[row.status] = result.get(row.status, 0) + 1
    return result


def count_by_test_status(rows: list[TestStatusRow]) -> dict[str, int]:
    result: dict[str, int] = {}
    for row in rows:
        result[row.status] = result.get(row.status, 0) + 1
    return result


def write_test_status_sheet(
    workbook: Workbook,
    used_names: set[str],
    rows: list[TestStatusRow],
) -> None:
    matrix = [test_status_headers()]
    matrix.extend(test_status_row_values(row) for row in rows)
    ws = workbook.create_sheet(unique_sheet_name(TEST_STATUS_SHEET, used_names))
    write_matrix(ws, matrix)
    style_status_table(ws, len(matrix), len(matrix[0]))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def test_status_headers() -> list[str]:
    return [
        "status",
        "group_order",
        "group_label",
        "source_sheet",
        "test_name",
        "baseline_distro",
        "compared_distro",
        "epsilon_percent",
        "metrics_total",
        "passed_metrics",
        "failed_metrics",
        "mean_value_delta_percent",
        "mean_performance_delta_percent",
        "failed_metric_names",
    ]


def test_status_row_values(row: TestStatusRow) -> list[Any]:
    return [
        row.status,
        row.group_order,
        row.group_label,
        row.source_sheet,
        row.test_name,
        row.baseline_distro,
        row.compared_distro,
        row.epsilon_percent,
        row.metrics_total,
        row.passed_metrics,
        row.failed_metrics,
        row.mean_value_delta_percent,
        row.mean_performance_delta_percent,
        row.failed_metric_names,
    ]


def write_metric_status_sheet(
    workbook: Workbook,
    used_names: set[str],
    rows: list[MetricStatusRow],
) -> None:
    matrix = [metric_status_headers()]
    matrix.extend(metric_status_row_values(row) for row in rows)
    ws = workbook.create_sheet(unique_sheet_name(METRIC_STATUS_SHEET, used_names))
    write_matrix(ws, matrix)
    style_status_table(ws, len(matrix), len(matrix[0]))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def metric_status_headers() -> list[str]:
    return [
        "status",
        "performance_change",
        "raw_value_change",
        "group_order",
        "group_label",
        "source_sheet",
        "test_name",
        "metric_name",
        "baseline_distro",
        "baseline_mean",
        "baseline_sample_count",
        "compared_distro",
        "compared_mean",
        "compared_sample_count",
        "better_direction",
        "epsilon_percent",
        "value_delta_percent",
        "performance_delta_percent",
    ]


def metric_status_row_values(row: MetricStatusRow) -> list[Any]:
    return [
        row.status,
        row.performance_change,
        row.raw_value_change,
        row.group_order,
        row.group_label,
        row.source_sheet,
        row.test_name,
        row.metric_name,
        row.baseline_distro,
        row.baseline_mean,
        row.baseline_sample_count,
        row.compared_distro,
        row.compared_mean,
        row.compared_sample_count,
        row.better_direction,
        row.epsilon_percent,
        row.value_delta_percent,
        row.performance_delta_percent,
    ]


def style_status_table(ws: Any, row_count: int, col_count: int) -> None:
    style_table(ws, row_count, col_count)
    status_fills = {
        STATUS_OK: PatternFill("solid", fgColor="D9EAD3"),
        STATUS_FAIL: PatternFill("solid", fgColor="F4CCCC"),
        STATUS_NOT_COMPARABLE: PatternFill("solid", fgColor="FFF2CC"),
    }
    for row_index in range(2, row_count + 1):
        status_cell = ws.cell(row=row_index, column=1)
        fill = status_fills.get(str(status_cell.value or ""))
        if fill is not None:
            status_cell.fill = fill

    for column_name in (
        "epsilon_percent",
        "value_delta_percent",
        "performance_delta_percent",
        "mean_value_delta_percent",
        "mean_performance_delta_percent",
    ):
        column_index_1based = worksheet_column_index(ws, column_name)
        if column_index_1based is None:
            continue
        for row_index in range(2, row_count + 1):
            ws.cell(row=row_index, column=column_index_1based).number_format = "0.00"


def worksheet_column_index(ws: Any, column_name: str) -> int | None:
    for cell in ws[1]:
        if cell.value == column_name:
            return cell.column
    return None
