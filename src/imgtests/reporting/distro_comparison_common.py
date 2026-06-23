from __future__ import annotations

import logging
import math
import re
from collections import defaultdict
from collections.abc import Callable, Iterator
from dataclasses import dataclass
from typing import TYPE_CHECKING, Any, Final, Literal, cast, get_args

from openpyxl.styles import Font, PatternFill

from imgtests.database.models.experiment import ExperimentType

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

DISTROS: Final = ("poky", "suse")
CONFIGURATION_SHEET: Final = "configuration"
EXPERIMENT_SHEET: Final = "experiment"

COMPARISON_SELECTION_SHEET: Final = "comparison_selection"
COMPARISON_INDEX_SHEET: Final = "comparison_index"
COMPARISON_DATA_SHEET: Final = "comparison_data"
COMPARISON_CHARTS_PREFIX: Final = "comparison_charts"
CHART_SHEET_PREFIX: Final = "charts"

NONZERO_EPSILON: Final = 1e-12
INVALID_NUMERIC_SENTINELS: Final = {-1.0}
SHORT_TOKEN_LENGTH: Final = 3
EXPLICIT_PAIR_EXPERIMENT_COUNT: Final = 2
MAX_SHEET_NAME_LENGTH: Final = 31
ID_COLUMN_SUFFIX: Final = "_id"
EXPERIMENT_TYPE_VALUES: Final = frozenset(get_args(ExperimentType))

SKIP_SOURCE_SHEETS: Final = {
    COMPARISON_SELECTION_SHEET,
    COMPARISON_INDEX_SHEET,
    COMPARISON_DATA_SHEET,
    COMPARISON_CHARTS_PREFIX,
    "comparison_metrics",
    "comparison_summary",
}

LABEL_COLUMNS: Final = {
    "configuration_id",
    "distribution_description",
    "experiment_id",
    "run_result_id",
    "id",
    "test_name",
    "tool_name",
    "util_type",
    "command",
    "description",
    "started_at",
    "ended_at",
    "type",
    "config_id",
    "tests_total",
    "tests_passed",
    "tests_failed",
    "tests_broken",
    "tests_skipped",
}

TECHNICAL_TOKENS: Final = {
    "id",
    "pid",
    "uid",
    "gid",
    "fd",
    "rc",
    "code",
    "exit",
    "status",
    "returncode",
    "return",
    "retval",
    "ret",
    "socket",
    "port",
    "host",
    "hostname",
    "ip",
    "addr",
    "address",
    "local",
    "remote",
    "timestamp",
    "epoch",
    "unix",
    "version",
    "command",
    "cmd",
    "cookie",
    "mss",
    "tos",
}

GENERATED_RESULTS_METRIC_RE: Final = re.compile(
    r"^results_[0-9a-f]{7,64}_results(?:_|$)",
)
CONFIGURATION_METRIC_NAMES: Final = {
    "target_bitrate",
    "target_bitrate_bps",
    "target_bitrate_bits_per_second",
    "test_duration",
    "test_duration_sec",
    "test_duration_secs",
    "test_duration_seconds",
    "timeout",
    "timeout_sec",
    "timeout_secs",
    "timeout_seconds",
}
SUFFIX_CONFIGURATION_METRIC_NAMES: Final = CONFIGURATION_METRIC_NAMES | {
    "bitrate_bps",
    "datagram_size_bytes",
    "jobs_count",
    "profiles_count",
    "start_offset_sec",
}
EXACT_CONFIGURATION_METRIC_NAMES: Final = SUFFIX_CONFIGURATION_METRIC_NAMES | {
    "iterations",
    "latency_depth",
    "latency_percentile",
    "latency_target",
    "latency_window",
    "pps",
    "total_iterations",
}
GENERIC_METRIC_NAMES: Final = {"value"}
RAW_OUTPUT_METRIC_PREFIXES: Final = ("stdout_",)

METRIC_TECHNICAL_TOKENS: Final = TECHNICAL_TOKENS | {
    "arg",
    "args",
    "argument",
    "arguments",
    "flag",
    "flags",
    "option",
    "options",
    "param",
    "parameter",
    "parameters",
    "config",
    "configuration",
    "compiler",
}

TIME_PATTERNS: Final = {
    "systemd",
    "systemd_analyze",
    "delay",
    "duration",
    "elapsed",
    "response",
    "response_time",
    "runtime",
    "sec",
    "seconds",
    "secs",
    "time",
    "latency",
    "lat",
    "clat",
    "slat",
    "rtt",
    "microseconds",
    "usec",
    "usecs",
    "usec_per_op",
    "usecs_per_op",
    "wait",
}

THROUGHPUT_PATTERNS: Final = {
    "bps",
    "bits_per_second",
    "bitrate",
    "bandwidth",
    "bw",
    "gb_per_sec",
    "mb_per_sec",
    "kb_per_sec",
    "iops",
    "ops_per_sec",
    "ops_s",
    "bogo_ops_per_sec",
    "bogo_ops_s",
    "per_sec",
    "pps",
    "requests_per_second",
    "records_per_second",
    "speed",
}

PERCENT_PATTERNS: Final = {
    "percent",
    "percentage",
    "pct",
    "ratio",
    "rate",
    "util",
    "usage",
    "cpu_used",
}

COUNT_PATTERNS: Final = {
    "count",
    "total",
    "num",
    "number",
    "items",
    "requests",
    "retransmit",
    "failed",
    "errors",
    "oom",
    "broken",
    "skipped",
}

MEMORY_PATTERNS: Final = {
    "disk_usage",
    "rss",
    "memory",
    "mem",
    "ram",
    "kb",
    "mb",
    "bytes",
}

IPERF_PATTERNS: Final = {
    "iperf",
    "network_loopback",
}

STRESS_NG_PATTERNS: Final = {
    "stress_ng",
}

MetricColumnSelector = Callable[[list[str]], list[int]]
MetricNameMapper = Callable[[str, str, str], str]
MetricPredicate = Callable[[str, str, str], bool]


@dataclass(frozen=True)
class SheetRows:
    name: str
    headers: list[str]
    rows: list[list[Any]]


@dataclass(frozen=True)
class ExperimentInfo:
    experiment_id: int
    distro: str
    configuration_id: int
    description: str
    experiment_type: ExperimentType | Literal[""]
    started_at: str


@dataclass(frozen=True)
class ComparisonGroup:
    order: int
    group_id: str
    label: str
    short_label: str
    experiments: dict[str, ExperimentInfo]


@dataclass
class MetricBucket:
    logical_test: str
    source_sheet: str
    metric_name: str
    values: dict[str, dict[str, list[float]]]


@dataclass(frozen=True)
class MetricWorksheetContext:
    sheet_name: str
    source_sheet: str
    rows: Iterator[tuple[Any, ...]]
    headers: list[str]
    config_index: int
    experiment_index: int
    metric_indexes: list[int]
    test_name_index: int | None
    tool_name_index: int | None
    util_type_index: int | None
    distribution_index: int | None


@dataclass(frozen=True)
class MetricExtractionOptions:
    column_selector: MetricColumnSelector
    name_mapper: MetricNameMapper | None = None
    predicate: MetricPredicate | None = None
    require_nonzero: bool = True


def is_comparison_sheet(sheet_name: str) -> bool:
    return sheet_name in SKIP_SOURCE_SHEETS or sheet_name.startswith(
        (f"{COMPARISON_CHARTS_PREFIX}_", f"{CHART_SHEET_PREFIX}_"),
    )


def build_configuration_distro_map(workbook: Workbook) -> dict[str, str]:
    if CONFIGURATION_SHEET not in workbook.sheetnames:
        return {}
    ws = workbook[CONFIGURATION_SHEET]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [normalize_header(value) for value in next(rows)]
    except StopIteration:
        return {}

    config_index = column_index(headers, "configuration_id")
    distro_index = column_index(headers, "distribution_description")
    if config_index is None or distro_index is None:
        return {}

    result: dict[str, str] = {}
    for row in rows:
        key = id_key(cell_value(row, config_index))
        distro = normalize_distro(cell_value(row, distro_index))
        if key and distro:
            result[key] = distro
    return result


def build_experiment_info_map(
    workbook: Workbook,
    config_distro: dict[str, str],
) -> dict[int, ExperimentInfo]:
    if EXPERIMENT_SHEET not in workbook.sheetnames:
        return {}
    ws = workbook[EXPERIMENT_SHEET]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [normalize_header(value) for value in next(rows)]
    except StopIteration:
        return {}

    config_index = column_index(headers, "configuration_id")
    experiment_index = column_index(headers, "experiment_id")
    description_index = column_index(headers, "description")
    type_index = column_index(headers, "type")
    started_at_index = column_index(headers, "started_at")
    if config_index is None or experiment_index is None:
        return {}

    result: dict[int, ExperimentInfo] = {}
    for row in rows:
        experiment_id = int_id_or_none(cell_value(row, experiment_index))
        configuration_id = int_id_or_none(cell_value(row, config_index))
        if experiment_id is None or configuration_id is None:
            continue
        result[experiment_id] = ExperimentInfo(
            experiment_id=experiment_id,
            distro=config_distro.get(id_key(configuration_id), ""),
            configuration_id=configuration_id,
            description=str(cell_value(row, description_index) or "").strip(),
            experiment_type=normalize_experiment_type(cell_value(row, type_index)),
            started_at=str(cell_value(row, started_at_index) or "").strip(),
        )
    return result


def build_comparison_groups(
    experiment_info: dict[int, ExperimentInfo],
    *,
    experiment_ids: list[str] | None,
) -> list[ComparisonGroup]:
    selected_ids = normalize_selected_ids(experiment_ids)
    if experiment_ids and not selected_ids:
        logger.warning("No valid integer experiment IDs were provided.")
        return []

    if selected_ids and len(selected_ids) == EXPLICIT_PAIR_EXPERIMENT_COUNT:
        ordered_ids = sorted(selected_ids)
        infos = [experiment_info.get(experiment_id) for experiment_id in ordered_ids]
        if not all(info is not None for info in infos):
            missing = [
                str(experiment_id)
                for experiment_id, info in zip(ordered_ids, infos, strict=True)
                if info is None
            ]
            logger.warning("Experiment IDs not found in experiment sheet: %s", ", ".join(missing))
            return []
        by_distro = {
            info.distro: info for info in infos if info is not None and info.distro in DISTROS
        }
        if not all(distro in by_distro for distro in DISTROS):
            logger.warning(
                "Two IDs were provided, but they are not one Poky and one SUSE experiment.",
            )
            return []
        return [make_group(by_distro["poky"], by_distro["suse"], order=1)]

    filtered_infos = [
        info
        for info in experiment_info.values()
        if info.distro in DISTROS and (not selected_ids or info.experiment_id in selected_ids)
    ]
    grouped: dict[tuple[str, str], dict[str, list[ExperimentInfo]]] = defaultdict(
        lambda: defaultdict(list),
    )
    for info in filtered_infos:
        grouped[(info.description, info.experiment_type)][info.distro].append(info)

    groups: list[ComparisonGroup] = []
    order = 1

    def experiment_sort_key(info: ExperimentInfo) -> tuple[str, int]:
        return info.started_at, info.experiment_id

    for key in sorted(grouped, key=lambda item: (item[0], item[1])):
        by_distro = grouped[key]
        if not all(distro in by_distro for distro in DISTROS):
            continue
        poky_items = sorted(by_distro["poky"], key=experiment_sort_key)
        suse_items = sorted(by_distro["suse"], key=experiment_sort_key)
        pair_count = min(len(poky_items), len(suse_items))
        if len(poky_items) != len(suse_items):
            logger.warning(
                "Experiment group '%s/%s' has %s Poky and %s SUSE runs; "
                "only %s comparable pairs will be used.",
                key[0],
                key[1],
                len(poky_items),
                len(suse_items),
                pair_count,
            )
        for poky_info, suse_info in zip(
            poky_items[:pair_count],
            suse_items[:pair_count],
            strict=True,
        ):
            groups.append(make_group(poky_info, suse_info, order=order))
            order += 1

    return groups


def make_group(
    poky_info: ExperimentInfo,
    suse_info: ExperimentInfo,
    *,
    order: int,
) -> ComparisonGroup:
    group_id = f"exp_{poky_info.experiment_id}_{suse_info.experiment_id}"
    label_parts = [part for part in (poky_info.description, poky_info.experiment_type) if part]
    base_label = " | ".join(label_parts) or "experiment"
    short_label = f"{order}: p{poky_info.experiment_id}/s{suse_info.experiment_id}"
    label = (
        f"{base_label} | exp {poky_info.experiment_id} / poky vs "
        f"exp {suse_info.experiment_id} / suse"
    )
    return ComparisonGroup(
        order=order,
        group_id=group_id,
        label=label,
        short_label=short_label,
        experiments={"poky": poky_info, "suse": suse_info},
    )


def normalize_selected_ids(experiment_ids: list[str] | None) -> set[int]:
    if not experiment_ids:
        return set()
    selected_ids: set[int] = set()
    invalid_ids: list[str] = []
    for experiment_id in experiment_ids:
        normalized_id = int_id_or_none(experiment_id)
        if normalized_id is None:
            invalid_ids.append(str(experiment_id))
            continue
        selected_ids.add(normalized_id)

    if invalid_ids:
        logger.warning("Experiment IDs must be integers and were ignored: %s", invalid_ids)
    return selected_ids


def build_group_by_experiment(groups: list[ComparisonGroup]) -> dict[str, ComparisonGroup]:
    group_by_experiment: dict[str, ComparisonGroup] = {}
    for group in groups:
        for info in group.experiments.values():
            group_by_experiment[id_key(info.experiment_id)] = group
    return group_by_experiment


def extract_metric_buckets(
    workbook: Workbook,
    config_distro: dict[str, str],
    groups: list[ComparisonGroup],
    options: MetricExtractionOptions,
) -> list[MetricBucket]:
    group_by_experiment = build_group_by_experiment(groups)
    buckets: dict[tuple[str, str, str], MetricBucket] = {}

    for worksheet in workbook.worksheets:
        context = metric_worksheet_context(worksheet, options.column_selector)
        if context is None:
            continue
        for row in context.rows:
            collect_metric_buckets_from_row(
                row,
                context,
                group_by_experiment,
                config_distro,
                buckets,
                options,
            )

    return list(buckets.values())


def metric_worksheet_context(
    worksheet: Worksheet,
    column_selector: MetricColumnSelector,
) -> MetricWorksheetContext | None:
    sheet_name = worksheet.title
    if is_comparison_sheet(sheet_name):
        return None

    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        headers = [normalize_header(value) for value in next(rows_iter)]
    except StopIteration:
        return None

    config_index = column_index(headers, "configuration_id")
    experiment_index = column_index(headers, "experiment_id")
    if config_index is None or experiment_index is None:
        return None

    metric_indexes = column_selector(headers)
    if not metric_indexes:
        return None

    return MetricWorksheetContext(
        sheet_name=sheet_name,
        source_sheet=strip_excel_suffix(sheet_name),
        rows=rows_iter,
        headers=headers,
        config_index=config_index,
        experiment_index=experiment_index,
        metric_indexes=metric_indexes,
        test_name_index=column_index(headers, "test_name"),
        tool_name_index=column_index(headers, "tool_name"),
        util_type_index=column_index(headers, "util_type"),
        distribution_index=column_index(headers, "distribution_description"),
    )


def collect_metric_buckets_from_row(  # noqa: PLR0913
    row: tuple[Any, ...],
    context: MetricWorksheetContext,
    group_by_experiment: dict[str, ComparisonGroup],
    config_distro: dict[str, str],
    buckets: dict[tuple[str, str, str], MetricBucket],
    options: MetricExtractionOptions,
) -> None:
    experiment_id = id_key(cell_value(row, context.experiment_index))
    group = group_by_experiment.get(experiment_id)
    if group is None:
        return

    distro = row_distro(row, context.distribution_index, context.config_index, config_distro)
    if distro not in DISTROS:
        return

    logical_test = logical_test_name(
        context.sheet_name,
        cell_value(row, context.test_name_index),
        cell_value(row, context.tool_name_index),
        cell_value(row, context.util_type_index),
    )

    for metric_index in context.metric_indexes:
        metric_name = context.headers[metric_index]
        bucket_metric_name = metric_name
        if options.name_mapper is not None:
            bucket_metric_name = options.name_mapper(
                context.source_sheet,
                logical_test,
                metric_name,
            )
        if options.predicate is not None and not options.predicate(
            context.source_sheet,
            logical_test,
            bucket_metric_name,
        ):
            continue
        value = to_finite_float(cell_value(row, metric_index))
        if value is None or is_invalid_metric_value(value):
            continue
        if options.require_nonzero and not is_nonzero_number(value):
            continue
        bucket = get_or_create_metric_bucket(
            buckets,
            logical_test,
            context.source_sheet,
            bucket_metric_name,
        )
        bucket.values[group.group_id][distro].append(value)


def get_or_create_metric_bucket(
    buckets: dict[tuple[str, str, str], MetricBucket],
    logical_test: str,
    source_sheet: str,
    metric_name: str,
) -> MetricBucket:
    bucket_key = (logical_test, source_sheet, metric_name)
    bucket = buckets.get(bucket_key)
    if bucket is None:
        bucket = MetricBucket(
            logical_test=logical_test,
            source_sheet=source_sheet,
            metric_name=metric_name,
            values=defaultdict(lambda: defaultdict(list)),
        )
        buckets[bucket_key] = bucket
    return bucket


def find_metric_columns(headers: list[str]) -> list[int]:
    metric_indexes: list[int] = []
    for index, header in enumerate(headers):
        if not header or header in LABEL_COLUMNS or header.endswith(ID_COLUMN_SUFFIX):
            continue
        normalized_header = normalize_metric_text(header)
        if not normalized_header:
            continue
        tokens = set(normalized_header.split("_"))
        if tokens & TECHNICAL_TOKENS:
            continue
        metric_indexes.append(index)
    return metric_indexes


def find_comparison_metric_columns(headers: list[str]) -> list[int]:
    metric_indexes: list[int] = []
    for index in find_metric_columns(headers):
        normalized_header = normalize_metric_text(headers[index])
        if is_ignored_metric_name(normalized_header):
            continue
        tokens = set(normalized_header.split("_"))
        if tokens & METRIC_TECHNICAL_TOKENS:
            continue
        metric_indexes.append(index)
    return metric_indexes


def is_ignored_metric_name(metric_name: str) -> bool:
    return (
        is_generated_results_metric_name(metric_name)
        or is_sample_count_metric_name(metric_name)
        or is_configuration_metric_name(metric_name)
        or is_raw_output_metric_name(metric_name)
        or is_generic_metric_name(metric_name)
    )


def is_generated_results_metric_name(metric_name: str) -> bool:
    return GENERATED_RESULTS_METRIC_RE.match(normalize_metric_text(metric_name)) is not None


def is_sample_count_metric_name(metric_name: str) -> bool:
    normalized = normalize_metric_text(metric_name)
    return normalized in {"n", "samples"} or normalized.endswith(("_n", "_samples"))


def is_configuration_metric_name(metric_name: str) -> bool:
    normalized = normalize_metric_text(metric_name)
    return normalized in EXACT_CONFIGURATION_METRIC_NAMES or any(
        normalized.endswith(f"_{name}") for name in SUFFIX_CONFIGURATION_METRIC_NAMES
    )


def is_raw_output_metric_name(metric_name: str) -> bool:
    return normalize_metric_text(metric_name).startswith(RAW_OUTPUT_METRIC_PREFIXES)


def is_generic_metric_name(metric_name: str) -> bool:
    return normalize_metric_text(metric_name) in GENERIC_METRIC_NAMES


def row_distro(
    row: tuple[Any, ...],
    distribution_index: int | None,
    config_index: int,
    config_distro: dict[str, str],
) -> str:
    if distribution_index is not None:
        distro = normalize_distro(cell_value(row, distribution_index))
        if distro:
            return distro
    return config_distro.get(id_key(cell_value(row, config_index)), "")


def logical_test_name(sheet_name: str, test_name: Any, tool_name: Any, util_type: Any) -> str:
    parts: list[str] = []
    prepared_test_name = str(test_name or "").strip()
    prepared_tool_name = str(tool_name or "").strip()
    prepared_util_type = str(util_type or "").strip()

    if prepared_test_name:
        parts.append(prepared_test_name)
    else:
        parts.append(strip_excel_suffix(sheet_name))

    if prepared_tool_name and prepared_tool_name.lower() not in prepared_test_name.lower():
        parts.append(prepared_tool_name)
    if prepared_util_type:
        parts.append(prepared_util_type)
    return " / ".join(parts)


def strip_excel_suffix(sheet_name: str) -> str:
    return re.sub(r"_\d+$", "", str(sheet_name).strip())


def normalize_header(value: Any) -> str:
    return "" if value is None else str(value).strip()


def column_index(headers: list[str], name: str) -> int | None:
    try:
        return headers.index(name)
    except ValueError:
        return None


def cell_value(row: tuple[Any, ...] | list[Any], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def row_has_data(row: tuple[Any, ...] | list[Any]) -> bool:
    return any(not is_empty_value(value) for value in row)


def sheet_has_data_rows(sheet: SheetRows) -> bool:
    return any(row_has_data(row) for row in sheet.rows)


def is_empty_value(value: Any) -> bool:
    return value is None or value == ""


def id_key(value: Any) -> str:
    if isinstance(value, bool) or value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def int_id_or_none(value: Any) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None

    text = str(value).strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def normalize_distro(value: Any) -> str:
    text = str(value or "").strip().lower()
    if "suse" in text:
        return "suse"
    if "yocto" in text or "poky" in text:
        return "poky"
    return text


def normalize_metric_text(value: str) -> str:
    normalized = re.sub(r"[^a-zA-Z0-9]+", "_", str(value).lower())
    normalized = re.sub(r"_+", "_", normalized)
    return normalized.strip("_")


def normalize_experiment_type(value: Any) -> ExperimentType | Literal[""]:
    text = str(value or "").strip()
    if text in EXPERIMENT_TYPE_VALUES:
        return cast("ExperimentType", text)
    return ""


def metric_matches(metric_text: str, metric_tokens: set[str], pattern: str) -> bool:
    if len(pattern) <= SHORT_TOKEN_LENGTH:
        return pattern in metric_tokens
    return pattern in metric_text


def metric_has_any_pattern(metric_text: str, metric_tokens: set[str], patterns: set[str]) -> bool:
    return any(metric_matches(metric_text, metric_tokens, pattern) for pattern in patterns)


def sortable_value(value: Any) -> tuple[int, int, str]:
    text = str(value or "").strip()
    try:
        return (0, int(text), "")
    except ValueError:
        return (1, 0, text)


def to_finite_float(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        converted = float(value)
        return converted if math.isfinite(converted) else None
    if isinstance(value, str):
        prepared = value.strip().replace(",", ".")
        if not prepared:
            return None
        try:
            converted = float(prepared)
        except ValueError:
            return None
        return converted if math.isfinite(converted) else None
    return None


def is_nonzero_number(value: float | None) -> bool:
    return value is not None and abs(value) > NONZERO_EPSILON


def is_invalid_metric_value(value: Any) -> bool:
    return value in INVALID_NUMERIC_SENTINELS


def resolve_output_path(input_path: Path, output_path: Path | None, output_name: str) -> Path:
    if output_path is None:
        return input_path.with_name(f"{input_path.stem}_{output_name}{input_path.suffix}")
    if output_path.suffix.lower() == ".xlsx":
        return output_path
    return output_path / f"{input_path.stem}_{output_name}.xlsx"


def write_matrix(ws: Worksheet, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(list(row))


def style_table(ws: Worksheet, row_count: int, col_count: int) -> None:
    if row_count <= 0 or col_count <= 0:
        return
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        width = min(max(len(str(cell.value or "")) + 2, 12), 45)
        ws.column_dimensions[cell.column_letter].width = width
    for row in ws.iter_rows(min_row=2, max_row=min(row_count, 200), max_col=col_count):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000"


def unique_sheet_name(sheet_name: str, used_names: set[str]) -> str:
    cleaned = sanitize_sheet_name(sheet_name)
    original = cleaned
    suffix = 1
    while cleaned in used_names:
        suffix_text = f"_{suffix}"
        cleaned = f"{original[: MAX_SHEET_NAME_LENGTH - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_names.add(cleaned)
    return cleaned


def sanitize_sheet_name(name: str) -> str:
    cleaned = "".join("_" if char in r":\/?*[]" else char for char in str(name)).strip("'")
    return (cleaned or "Sheet")[:MAX_SHEET_NAME_LENGTH]
